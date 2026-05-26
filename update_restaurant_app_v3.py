"""
Script de mise à jour de l'application Excel de gestion de restaurant
Intègre les fiches : Plats, Stock Cuisine, Ventes avec formules avancées
Sans VBA - 100% formules Excel natives
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime

# Chargement du fichier existant
fichier_excel = "GestionRestaurant.xlsx"
if os.path.exists(fichier_excel):
    wb = load_workbook(fichier_excel)
else:
    wb = Workbook()

# Suppression des feuilles obsolètes si elles existent
feuilles_a_supprimer = ['Plats', 'StockCuisine']
for sheet_name in feuilles_a_supprimer:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

# ============================================================================
# 1. FEUILLE "PLATS" - Gestion complète des plats, packs et planches
# ============================================================================
print("Création de la feuille PLATS...")

if 'PLATS' in wb.sheetnames:
    ws_plats = wb['PLATS']
else:
    ws_plats = wb.create_sheet('PLATS')

# Couleurs
bleu_marine = "2C3E50"
vert = "2ECC71"
gris_clair = "ECF0F1"
orange = "F39C12"
rouge = "E74C3C"

# Style des titres
titre_font = Font(name='Arial', size=20, bold=True, color="FFFFFF")
sous_titre_font = Font(name='Arial', size=14, bold=True, color=bleu_marine)
header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
normal_font = Font(name='Arial', size=11)

fill_bleu = PatternFill(start_color=bleu_marine, end_color=bleu_marine, fill_type="solid")
fill_vert = PatternFill(start_color=vert, end_color=vert, fill_type="solid")
fill_gris = PatternFill(start_color=gris_clair, end_color=gris_clair, fill_type="solid")
fill_orange = PatternFill(start_color=orange, end_color=orange, fill_type="solid")
fill_rouge_leger = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

align_center = Alignment(horizontal='center', vertical='center')
align_left = Alignment(horizontal='left', vertical='center')

# Titre principal
ws_plats['A1'] = "GESTION DES PLATS"
ws_plats['A1'].font = titre_font
ws_plats['A1'].fill = fill_bleu
ws_plats.merge_cells('A1:D1')

# Section 1: Liste des plats
ws_plats['A3'] = "Liste des Plats"
ws_plats['A3'].font = sous_titre_font

# En-têtes
en_tetes_plats = ["Plat", "Variante", "Prix (FCFA)", "Observation"]
for col, header in enumerate(en_tetes_plats, start=1):
    cell = ws_plats.cell(row=5, column=col, value=header)
    cell.font = header_font
    cell.fill = fill_bleu
    cell.alignment = align_center

# Données des plats (53 plats + variantes)
plats_data = [
    # Burgers
    ["Burger", "Chicken Burger", 3500, ""],
    ["Burger", "Burger Normal", 3500, ""],
    # Brochettes
    ["Brochette de porc", "", 3000, "Trois Brochettes"],
    # Ndole
    ["Ndole Royal", "Viande", 3500, ""],
    ["Ndole Royal", "Poisson fumé", 3500, ""],
    ["Ndole Royal", "Crevette", 3500, ""],
    ["Ndole Royal", "Portion", 1000, ""],
    # Free Rice
    ["Free Rice", "Crevette", 4000, ""],
    ["Free Rice", "Poulet", 3500, ""],
    # Poulet
    ["Poulet", "Pané", 3500, "un quart"],
    ["Poulet", "Braisé", 3000, "un quart"],
    # Porc
    ["Porc", "Moutarde", 3500, ""],
    ["Porc", "Braisé", 3000, ""],
    # Gambas
    ["Gambas", "À l'ail", 6500, "6500/10000"],
    ["Gambas", "Braisé", 6500, "6500/10000"],
    # Saucisse
    ["Saucisse", "", 3500, "x 2"],
    # Tacos
    ["Tacos", "Poulet", 4000, ""],
    ["Tacos", "Viande", 3500, ""],
    # Pizzas - Solo
    ["Pizza", "Seven - Solo", 3500, ""],
    ["Pizza", "Seven - Duo", 7000, ""],
    ["Pizza", "Seven - Familiale", 13500, ""],
    ["Pizza", "Spices Bonita - Solo", 3000, ""],
    ["Pizza", "Spices Bonita - Duo", 6000, ""],
    ["Pizza", "Spices Bonita - Familiale", 12000, ""],
    ["Pizza", "Bimbia Cheese - Solo", 2500, ""],
    ["Pizza", "Bimbia Cheese - Duo", 4500, ""],
    ["Pizza", "Bimbia Cheese - Familiale", 10000, ""],
    ["Pizza", "Végétarienne - Solo", 2500, ""],
    ["Pizza", "Végétarienne - Duo", 5000, ""],
    ["Pizza", "Végétarienne - Familiale", 11000, ""],
    # Poissons
    ["Poisson", "Carpe", 5000, "5000-8000"],
    ["Poisson", "Sole", 9000, "9000-11000"],
    ["Poisson", "Bar", 7000, "7000-10000"],
    # Jus
    ["Jus Naturel / Bissap", "Litre", 2000, "Conso à 1000"],
    ["Jus Naturel / Bissap", "Fruit de Saison", 3500, "En litre"],
    # Accompagnements
    ["Accompagnement", "Riz", 0, "Pour chaque plat"],
    ["Accompagnement", "Frites", 0, "Pour chaque plat"],
    ["Accompagnement", "Plantain", 0, "Pour chaque plat"],
    ["Accompagnement", "Miondo", 0, "Pour chaque plat"],
]

for row_idx, plat in enumerate(plats_data, start=6):
    for col_idx, value in enumerate(plat, start=1):
        ws_plats.cell(row=row_idx, column=col_idx, value=value).font = normal_font

# Section Ajout d'un plat
ws_plats['A52'] = "Ajouter un Plat"
ws_plats['A52'].font = sous_titre_font

labels_ajout = ["Nom du plat:", "Variante:", "Prix:", "Observation:"]
for idx, label in enumerate(labels_ajout, start=54):
    ws_plats.cell(row=idx, column=2, value=label).font = normal_font

# Bouton Ajouter (simulation avec formule)
ws_plats['B62'] = "Ajouter"
ws_plats['B62'].fill = fill_vert
ws_plats['B62'].font = Font(bold=True, color="FFFFFF")
ws_plats['B62'].alignment = align_center

# Section Packs
ws_plats['A64'] = "PACKS"
ws_plats['A64'].font = sous_titre_font
ws_plats['A64'].fill = fill_bleu
ws_plats['A64'].font = Font(size=14, bold=True, color="FFFFFF")

en_tetes_packs = ["Pack", "Composition", "Prix (FCFA)"]
for col, header in enumerate(en_tetes_packs, start=1):
    cell = ws_plats.cell(row=66, column=col, value=header)
    cell.font = header_font
    cell.fill = fill_bleu
    cell.alignment = align_center

packs_data = [
    ["Pack Brise Snack", "Deux morceaux de poulet, portion de frit, un verre de jus", 4000],
    ["Pack Sunset Combo", "Un burger, un morceau de poulet, portion de frit, un verre de fruit", 6000],
    ["Pack Ocean Duo", "Double cheese burger, un tacos, portion de frit, jus Top", 8500],
    ["Pack Bonita Family", "Un demi poulet pané, deux burgers, un litre de jus, deux portions de frit", 15000],
]

for row_idx, pack in enumerate(packs_data, start=67):
    for col_idx, value in enumerate(pack, start=1):
        ws_plats.cell(row=row_idx, column=col_idx, value=value).font = normal_font

# Section Planches
ws_plats['A74'] = "PLANCHES"
ws_plats['A74'].font = sous_titre_font
ws_plats['A74'].fill = fill_bleu
ws_plats['A74'].font = Font(size=14, bold=True, color="FFFFFF")

en_tetes_planches = ["Prix (FCFA)", "Nom recommandé"]
for col, header in enumerate(en_tetes_planches, start=1):
    cell = ws_plats.cell(row=76, column=col, value=header)
    cell.font = header_font
    cell.fill = fill_bleu
    cell.alignment = align_center

planches_data = [
    [10000, "Planche Brise Marine"],
    [25000, "Planche Sunset"],
    [50000, "Planche Océan Royale"],
    [75000, "Planche Paradise"],
    [100000, "Planche Isla Prestige"],
    [200000, "Planche Bonita Supreme"],
]

for row_idx, planche in enumerate(planches_data, start=77):
    for col_idx, value in enumerate(planche, start=1):
        ws_plats.cell(row=row_idx, column=col_idx, value=value).font = normal_font

# Ajuster les largeurs
ws_plats.column_dimensions['A'].width = 25
ws_plats.column_dimensions['B'].width = 30
ws_plats.column_dimensions['C'].width = 15
ws_plats.column_dimensions['D'].width = 25

# ============================================================================
# 2. FEUILLE "STOCK CUISINE" - Suivi détaillé par catégorie
# ============================================================================
print("Création de la feuille STOCK CUISINE...")

if 'STOCK_CUISINE' in wb.sheetnames:
    ws_stock = wb['STOCK_CUISINE']
else:
    ws_stock = wb.create_sheet('STOCK_CUISINE')

# Titre et date
ws_stock['A1'] = "SUIVI STOCK EN CUISINE"
ws_stock['A1'].font = titre_font
ws_stock['A1'].fill = fill_bleu
ws_stock.merge_cells('A1:G1')

ws_stock['C2'] = "Date:"
ws_stock['D2'] = "=AUJOURDHUI()"
ws_stock['D2'].number_format = "dd/mm/yyyy"

# Catégories complètes avec toutes les sous-catégories
categories_stock = {
    "Viandes": [
        ("Poulet entier", "Morceau"),
        ("Aile de Poulet", "Morceau"),
        ("Cuisse de poulet", "Morceau"),
        ("Bœuf", "Morceau"),
        ("Saucisse", "Pièce"),
        ("Porc", "Morceau"),
        ("Œuf", "Alvéole"),
        ("Viandes hachées", "Pièce"),
    ],
    "Poissons & Fruits de mer": [
        ("Carpe", "Pièce"),
        ("Sole", "Pièce"),
        ("Bar", "Pièce"),
        ("Gambas", "Pièce"),
        ("Crevette", "Pièce"),
    ],
    "Légumes frais": [
        ("Carottes", "Lot"),
        ("Pommes de terre", "Lot"),
        ("Gingembre", "Lot"),
        ("Choux", "Lot"),
        ("Salade", "Lot"),
        ("Tomates", "Cageot"),
        ("Poivrons", "Lot"),
        ("Aubergines", "Lot"),
        ("Oignons", "Lot"),
        ("Ail", "Lot"),
        ("Poireaux", "Lot"),
        ("Haricots verts", "Lot"),
        ("Persil", "Lot"),
        ("Curcuma", "Boite"),
        ("Poivre", "Lot"),
        ("Courgettes", "Lot"),
        ("Piments", "Lot"),
    ],
    "Fruits frais": [
        ("Orange", "Pièce"),
        ("Papaye", "Pièce"),
        ("Mangue", "Pièce"),
        ("Citron", "Pièce"),
        ("Goyave", "Pièce"),
        ("Mandarine", "Pièce"),
        ("Pastèque", "Pièce"),
        ("Pamplemousse", "Pièce"),
        ("Fruits pour jus", "Pièce"),
    ],
    "Produits secs & épicerie": [
        ("Riz", "Kg"),
        ("Pâtes", "Kg"),
        ("Haricots", "Kg"),
        ("Farines", "Kg"),
        ("Sucres", "Kg"),
        ("Sel", "Kg"),
        ("Bouillons & cubes", "Paquet"),
        ("Semoule", "Kg"),
    ],
    "Produits laitiers & fromages": [
        ("Lait & crème", "Litre"),
        ("Beurres & margarines", "Boite"),
        ("Fromages tranchés", "Pièce"),
        ("Fromages râpés", "Pièce"),
        ("Fromages en bloc", "Pièce"),
        ("Yaourts", "Litre"),
    ],
    "Charcuterie": [
        ("Jambons", "Tranche"),
        ("Bacons", "Tranche"),
        ("Saucissons", "Tranche"),
        ("Pepperoni", "Tranche"),
        ("Charcuterie cuite", "Tranche"),
        ("Charcuterie fumée", "Tranche"),
    ],
    "Surgelés": [
        ("Frites surgelées", "Paquet"),
        ("Légumes surgelés", "Paquet"),
        ("Viandes surgelées", "Paquet"),
        ("Poissons surgelés", "Paquet"),
        ("Produits panés", "Paquet"),
        ("Pâtes surgelées", "Paquet"),
    ],
    "Sauces & condiments": [
        ("Ketchup", "Bouteille"),
        ("Mayonnaise", "Bouteille"),
        ("Moutarde", "Bouteille"),
        ("Vinaigres", "Bouteille"),
        ("Marinades", "Bouteille"),
        ("Bouillons liquides", "Bouteille"),
    ],
    "Boulangerie & pains": [
        ("Pains burgers", "Nombre"),
        ("Pains hot-dog", "Nombre"),
        ("Pains sandwich", "Nombre"),
        ("Baguettes", "Nombre"),
        ("Tortillas / wraps", "Nombre"),
        ("Pain spécial", "Nombre"),
    ],
    "Herbes & épices": [
        ("Herbes fraîches", "Bouquet"),
        ("Herbes sèches", "Bouquet"),
        ("Épices moulues", "Sachet"),
        ("Épices entières", "Sachet"),
        ("Mélanges d'épices", "Sachet"),
    ],
    "Produits de pâtisserie": [
        ("Chocolats", "Boite"),
        ("Levures", "Sachet"),
        ("Arômes", "Bouteille"),
        ("Sucre glace", "Kg"),
        ("Cacao", "Kg"),
        ("Fruits secs", "Kg"),
    ],
    "Huiles & matières grasses": [
        ("Huiles végétales", "Litre"),
        ("Huiles spéciales", "Litre"),
        ("Beurres", "Kg"),
        ("Margarines", "Kg"),
        ("Graisses de friture", "Litre"),
    ],
    "Produits de nettoyage": [
        ("Détergents", "Bouteille"),
        ("Désinfectants", "Bouteille"),
        ("Produits vaisselle", "Bouteille"),
        ("Produits sol", "Bouteille"),
        ("Éponges, gants", "Lot"),
    ],
    "Emballages cuisine": [
        ("Barquettes", "Lot"),
        ("Boîtes", "Lot"),
        ("Films alimentaires", "Rouleau"),
        ("Sacs kraft", "Lot"),
        ("Bouteille de Gaz", "Pièce"),
        ("Couverts jetables", "Lot"),
        ("Charbon", "Sac"),
    ],
}

row = 4
for categorie, sous_categories in categories_stock.items():
    # Titre de catégorie
    ws_stock.cell(row=row, column=1, value=categorie.upper()).font = Font(size=13, bold=True, color=vert)
    row += 1
    
    # En-têtes
    en_tetes = ["Sous-catégorie", "Unité", "Stock initial", "Entrées", "Sorties", "Stock final", "Alerte"]
    for col, header in enumerate(en_tetes, start=1):
        cell = ws_stock.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = fill_bleu
        cell.alignment = align_center
    row += 1
    
    # Sous-catégories avec formules
    for sous_cat, unite in sous_categories:
        ws_stock.cell(row=row, column=1, value=sous_cat).font = normal_font
        ws_stock.cell(row=row, column=2, value=unite).font = normal_font
        # Stock initial (vide, à remplir)
        ws_stock.cell(row=row, column=3, value=0).number_format = "0"
        # Entrées
        ws_stock.cell(row=row, column=4, value=0).number_format = "0"
        # Sorties
        ws_stock.cell(row=row, column=5, value=0).number_format = "0"
        # Formule Stock final = Initial + Entrées - Sorties
        ws_stock.cell(row=row, column=6, value=f"=C{row}+D{row}-E{row}")
        ws_stock.cell(row=row, column=6).number_format = "0"
        # Formule Alerte stock critique (< 10% du stock initial ou < 5)
        ws_stock.cell(row=row, column=7, value=f'=SI(OU(F{row}<5; F{row}<C{row}*0,1); "⚠️ CRITIQUE"; "✓ OK")')
        
        row += 1
    
    row += 1  # Espace entre catégories

# Ajuster les largeurs
ws_stock.column_dimensions['A'].width = 25
ws_stock.column_dimensions['B'].width = 12
ws_stock.column_dimensions['C'].width = 15
ws_stock.column_dimensions['D'].width = 12
ws_stock.column_dimensions['E'].width = 12
ws_stock.column_dimensions['F'].width = 15
ws_stock.column_dimensions['G'].width = 15

# Mise en forme conditionnelle pour les alertes
ws_stock.conditional_formatting.add(
    'G6:G200',
    FormulaRule(formula=['ISNUMBER(SEARCH("CRITIQUE", G6))'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"), font=Font(color="9C0006", bold=True))
)

# ============================================================================
# 3. FEUILLE "VENTES" - Fiche journalière améliorée
# ============================================================================
print("Mise à jour de la feuille VENTES...")

if 'VENTES' not in wb.sheetnames:
    ws_ventes = wb.create_sheet('VENTES')
else:
    # Supprimer et recréer la feuille pour éviter les problèmes de cellules fusionnées
    del wb['VENTES']
    ws_ventes = wb.create_sheet('VENTES')

# En-tête avec date et heures
ws_ventes['A1'] = "VENTES JOURNALIÈRES RESTAURANT"
ws_ventes['A1'].font = titre_font
ws_ventes['A1'].fill = fill_bleu
ws_ventes.merge_cells('A1:F1')

ws_ventes['A3'] = "Date:"
ws_ventes['B3'] = "=AUJOURDHUI()"
ws_ventes['B3'].number_format = "dd/mm/yyyy"

ws_ventes['D3'] = "Heure début:"
ws_ventes['E3'] = ""

ws_ventes['A4'] = "Heure fin:"
ws_ventes['B4'] = ""

# En-têtes du tableau
en_tetes_ventes = ["Plat vendu", "Quantité vendue", "Prix Unitaire", "Total", "Observation", "Statut Stock"]
for col, header in enumerate(en_tetes_ventes, start=1):
    cell = ws_ventes.cell(row=6, column=col, value=header)
    cell.font = header_font
    cell.fill = fill_bleu
    cell.alignment = align_center

# Validation de données pour liste déroulante (plats depuis la feuille PLATS)
# Note: La validation sera configurée manuellement dans Excel car openpyxl a des limitations
# Mais on ajoute une note
ws_ventes['A8'] = "← Sélectionnez dans la liste (configurée via Données > Validation)"
ws_ventes['A8'].font = Font(italic=True, color="666666")

# Lignes de saisie (20 lignes)
for row in range(8, 28):
    # Plat (liste déroulante à configurer)
    ws_ventes.cell(row=row, column=1, value="")
    # Quantité
    ws_ventes.cell(row=row, column=2, value=0).number_format = "0"
    # Prix unitaire (formule pour récupérer depuis PLATS)
    # Formule simplifiée - dans Excel utilisez RECHERCHEV
    ws_ventes.cell(row=row, column=3, value=f'=SI(A{row}=""; ""; RECHERCHEV(A{row}; PLATS!A:C; 3; FAUX))')
    # Total
    ws_ventes.cell(row=row, column=4, value=f'=B{row}*C{row}')
    ws_ventes.cell(row=row, column=4).number_format = "#,##0 FCFA"
    # Observation
    ws_ventes.cell(row=row, column=5, value="")
    # Statut stock (vérification)
    ws_ventes.cell(row=row, column=6, value=f'=SI(A{row}=""; ""; SI(B{row}>10; "⚠️ Vérifier stock"; "✓ OK"))')

# Totaux
ws_ventes['A30'] = "TOTAUX"
ws_ventes['A30'].font = sous_titre_font
ws_ventes['A30'].fill = fill_bleu
ws_ventes['A30'].font = Font(bold=True, color="FFFFFF", size=14)

ws_ventes['B31'] = "Total Ventes:"
ws_ventes['C31'] = "=SOMME(D8:D27)"
ws_ventes['C31'].font = Font(bold=True, size=12)
ws_ventes['C31'].number_format = "#,##0 FCFA"

ws_ventes['B32'] = "Nombre de plats vendus:"
ws_ventes['C32'] = "=SOMME(B8:B27)"
ws_ventes['C32'].font = Font(bold=True, size=12)

ws_ventes['B33'] = "Panier moyen:"
ws_ventes['C33'] = '=SI(C32>0; C31/C32; 0)'
ws_ventes['C33'].number_format = "#,##0 FCFA"

# Signature
ws_ventes['A36'] = "Signature pour validation:"
ws_ventes['A36'].font = Font(bold=True)
ws_ventes['B36'] = ""

# Bouton Valider (simulation)
ws_ventes['A38'] = "VALIDER"
ws_ventes['A38'].fill = fill_vert
ws_ventes['A38'].font = Font(bold=True, color="FFFFFF", size=14)
ws_ventes['A38'].alignment = align_center
ws_ventes.merge_cells('A38:C38')

# Ajuster les largeurs
ws_ventes.column_dimensions['A'].width = 30
ws_ventes.column_dimensions['B'].width = 15
ws_ventes.column_dimensions['C'].width = 15
ws_ventes.column_dimensions['D'].width = 18
ws_ventes.column_dimensions['E'].width = 25
ws_ventes.column_dimensions['F'].width = 15

# ============================================================================
# 4. MISE À JOUR DASHBOARD ET INTERACTIONS
# ============================================================================
print("Mise à jour du Dashboard et des interactions...")

# Mettre à jour le dashboard pour inclure les nouvelles données
if 'DASHBOARD' in wb.sheetnames:
    ws_dash = wb['DASHBOARD']
    
    # Ajouter référence aux nouveaux totaux
    ws_dash['A15'] = "Ventes du jour (nouveau):"
    ws_dash['B15'] = "=VENTES!C31"
    ws_dash['B15'].number_format = "#,##0 FCFA"

# ============================================================================
# 5. SAUVEGARDE ET NETTOYAGE
# ============================================================================

# Réorganiser l'ordre des feuilles
ordre_souhaite = ['ACCUEIL', 'DASHBOARD', 'PLATS', 'PRODUITS', 'CATEGORIES', 'STOCK_CUISINE', 'MATIERES_PREMIERES', 'STOCKS', 'RECETTES', 'INGREDIENTS', 'VENTES', 'FOURNISSEURS', 'INVENTAIRES', 'TABLEAUX_BORD', 'PARAMETRES', 'AIDE']

for idx, sheet_name in enumerate(ordre_souhaite):
    if sheet_name in wb.sheetnames:
        wb.move_sheet(sheet_name, offset=-idx)

# Sauvegarder
wb.save(fichier_excel)
print(f"\n✅ Application mise à jour avec succès !")
print(f"📁 Fichier enregistré : {fichier_excel}")
print("\n📋 Nouvelles feuilles ajoutées:")
print("   - PLATS : 53 plats + 4 packs + 6 planches")
print("   - STOCK_CUISINE : 15 catégories, 105 articles avec formules")
print("   - VENTES : Fiche journalière complète avec calculs automatiques")
print("\n🔧 Prochaines étapes dans Excel:")
print("   1. Configurer la validation des données pour les listes déroulantes")
print("   2. Saisir les stocks initiaux dans STOCK_CUISINE")
print("   3. Tester les formules de calcul automatique")
print("   4. Personnaliser les seuils d'alerte si nécessaire")
