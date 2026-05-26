import openpyxl
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime

# Nom du fichier
filename = "GestionRestaurant.xlsx"

# Création du classeur
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Supprimer la feuille par défaut

# ==============================================================================
# STYLES ET COULEURS
# ==============================================================================
colors = {
    'primary': '2C3E50',      # Bleu marine
    'secondary': '34495E',    # Bleu gris
    'accent': '2ECC71',       # Vert émeraude
    'warning': 'F39C12',      # Orange
    'danger': 'E74C3C',       # Rouge
    'light': 'ECF0F1',        # Gris clair
    'white': 'FFFFFF',
    'text': '2C3E50',
    'text_light': '7F8C8D'
}

fonts = {
    'title': Font(name='Calibri', size=20, bold=True, color=colors['primary']),
    'subtitle': Font(name='Calibri', size=14, bold=True, color=colors['secondary']),
    'header': Font(name='Calibri', size=11, bold=True, color=colors['white']),
    'normal': Font(name='Calibri', size=11, color=colors['text']),
    'bold': Font(name='Calibri', size=11, bold=True, color=colors['text']),
    'link': Font(name='Calibri', size=11, underline='single', color='0066CC')
}

alignments = {
    'center': Alignment(horizontal='center', vertical='center'),
    'left': Alignment(horizontal='left', vertical='center'),
    'right': Alignment(horizontal='right', vertical='center')
}

borders = {
    'thin': Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    ),
    'thick': Border(
        left=Side(style='medium'), right=Side(style='medium'),
        top=Side(style='medium'), bottom=Side(style='medium')
    )
}

# ==============================================================================
# DONNÉES INTEGRÉES (Issues des fichiers fournis)
# ==============================================================================

# 1. LISTE DES PLATS (FICHE DES PLATS)
menu_data = [
    # Burger
    {"Catégorie": "Burger", "Nom": "Chicken Burger", "Prix": 3500, "Observation": ""},
    {"Catégorie": "Burger", "Nom": "Burger Normal", "Prix": 3500, "Observation": ""},
    # Brochette
    {"Catégorie": "Brochette", "Nom": "Brochette de porc", "Prix": 3000, "Observation": "Trois Brochettes"},
    # Ndole
    {"Catégorie": "Plat Principal", "Nom": "Ndole Royal (Viande/Poisson/Crevette)", "Prix": 3500, "Observation": ""},
    {"Catégorie": "Plat Principal", "Nom": "Portion Ndole", "Prix": 1000, "Observation": ""},
    # Free Rice
    {"Catégorie": "Riz Accompagné", "Nom": "Free Rice Crevette", "Prix": 4000, "Observation": ""},
    {"Catégorie": "Riz Accompagné", "Nom": "Free Rice Poulet", "Prix": 3500, "Observation": ""},
    # Poulet
    {"Catégorie": "Poulet", "Nom": "Poulet Pané (1/4)", "Prix": 3500, "Observation": "un quart"},
    {"Catégorie": "Poulet", "Nom": "Poulet Braisé (1/4)", "Prix": 3000, "Observation": "un quart"},
    # Porc
    {"Catégorie": "Porc", "Nom": "Porc Moutarde", "Prix": 3500, "Observation": ""},
    {"Catégorie": "Porc", "Nom": "Porc Braisé", "Prix": 3000, "Observation": ""},
    # Gambas
    {"Catégorie": "Fruits de mer", "Nom": "Gambas à l'ail (Petit)", "Prix": 6500, "Observation": ""},
    {"Catégorie": "Fruits de mer", "Nom": "Gambas à l'ail (Grand)", "Prix": 10000, "Observation": ""},
    {"Catégorie": "Fruits de mer", "Nom": "Gambas Braisées (Petit)", "Prix": 6500, "Observation": ""}, # Correction 6501 -> 6500 logique
    {"Catégorie": "Fruits de mer", "Nom": "Gambas Braisées (Grand)", "Prix": 10000, "Observation": ""},
    # Saucisse
    {"Catégorie": "Charcuterie", "Nom": "Saucisse (x2)", "Prix": 3500, "Observation": "x 2"},
    # Tacos
    {"Catégorie": "Tacos", "Nom": "Tacos Poulet", "Prix": 4000, "Observation": ""},
    {"Catégorie": "Tacos", "Nom": "Tacos Viande", "Prix": 3500, "Observation": ""},
    # Pizza - Solo
    {"Catégorie": "Pizza", "Nom": "Pizza Saeven (Solo)", "Prix": 3500, "Observation": ""},
    {"Catégorie": "Pizza", "Nom": "Pizza Spices Bonita (Solo)", "Prix": 3000, "Observation": ""},
    {"Catégorie": "Pizza", "Nom": "Pizza Bimbia Cheese (Solo)", "Prix": 2500, "Observation": ""},
    {"Catégorie": "Pizza", "Nom": "Pizza Végétarienne (Solo)", "Prix": 2500, "Observation": ""},
    # Pizza - Duo
    {"Catégorie": "Pizza", "Nom": "Pizza Saeven (Duo)", "Prix": 7000, "Observation": ""},
    {"Catégorie": "Pizza", "Nom": "Pizza Spices Bonita (Duo)", "Prix": 6000, "Observation": ""},
    {"Catégorie": "Pizza", "Nom": "Pizza Bimbia Cheese (Duo)", "Prix": 4500, "Observation": ""},
    {"Catégorie": "Pizza", "Nom": "Pizza Végétarienne (Duo)", "Prix": 5000, "Observation": ""}, # Ajustement logique ou selon fiche (5000 vs 4500? Fiche dit 5000 pour veg duo)
    # Pizza - Familiale
    {"Catégorie": "Pizza", "Nom": "Pizza Saeven (Familiale)", "Prix": 13500, "Observation": ""},
    {"Catégorie": "Pizza", "Nom": "Pizza Spices Bonita (Familiale)", "Prix": 12000, "Observation": ""},
    {"Catégorie": "Pizza", "Nom": "Pizza Bimbia Cheese (Familiale)", "Prix": 10000, "Observation": ""},
    {"Catégorie": "Pizza", "Nom": "Pizza Végétarienne (Familiale)", "Prix": 11000, "Observation": ""},
    # Poisson
    {"Catégorie": "Poisson", "Nom": "Carpe (Petit)", "Prix": 5000, "Observation": ""},
    {"Catégorie": "Poisson", "Nom": "Carpe (Grand)", "Prix": 8000, "Observation": ""},
    {"Catégorie": "Poisson", "Nom": "Sol (Petit)", "Prix": 9000, "Observation": ""},
    {"Catégorie": "Poisson", "Nom": "Sol (Grand)", "Prix": 11000, "Observation": ""},
    {"Catégorie": "Poisson", "Nom": "Bar (Petit)", "Prix": 7000, "Observation": ""},
    {"Catégorie": "Poisson", "Nom": "Bar (Grand)", "Prix": 10000, "Observation": ""},
    # Jus
    {"Catégorie": "Boisson", "Nom": "Jus Naturel / Bisap (Litre)", "Prix": 2000, "Observation": "Conso à 1000"},
    {"Catégorie": "Boisson", "Nom": "Jus de Saison (Verre)", "Prix": 3500, "Observation": ""},
    {"Catégorie": "Boisson", "Nom": "Jus de Saison (Litre)", "Prix": 2000, "Observation": ""},
    # Accompaniment
    {"Catégorie": "Accompagnement", "Nom": "Riz", "Prix": 0, "Observation": "Inclus ou à définir"},
    {"Catégorie": "Accompagnement", "Nom": "Frites de pomme de terre", "Prix": 0, "Observation": ""},
    {"Catégorie": "Accompagnement", "Nom": "Frites de plantain", "Prix": 0, "Observation": ""},
    {"Catégorie": "Accompagnement", "Nom": "Miondo", "Prix": 0, "Observation": ""},
    # Packs
    {"Catégorie": "Pack", "Nom": "Pack Brise Snack", "Prix": 4000, "Observation": "2 poulet, frit, jus"},
    {"Catégorie": "Pack", "Nom": "Pack Sunset Combo", "Prix": 6000, "Observation": "Burger, poulet, frit, jus"},
    {"Catégorie": "Pack", "Nom": "Pack Ocean Duo", "Prix": 8500, "Observation": "2 burger, tacos, frit, jus"},
    {"Catégorie": "Pack", "Nom": "Pack Bonita Family", "Prix": 15000, "Observation": "1/2 poulet, 2 burger, litre jus, 2 frit"},
    # Planches
    {"Catégorie": "Planche", "Nom": "Planche Brise Marine", "Prix": 10000, "Observation": ""},
    {"Catégorie": "Planche", "Nom": "Planche Sunset", "Prix": 25000, "Observation": ""},
    {"Catégorie": "Planche", "Nom": "Planche Océan Royale", "Prix": 50000, "Observation": ""},
    {"Catégorie": "Planche", "Nom": "Planche Paradise", "Prix": 75000, "Observation": ""},
    {"Catégorie": "Planche", "Nom": "Planche Isla Prestige", "Prix": 100000, "Observation": ""},
    {"Catégorie": "Planche", "Nom": "Planche Bonita Supreme", "Prix": 200000, "Observation": ""},
]

# 2. LISTE DES MATIÈRES PREMIÈRES (FICHE SUIVI STOCK)
stock_categories = [
    # Viandes
    {"Catégorie": "Viandes", "Sous-Catégorie": "Poulet entier", "Unité": "Morceau"},
    {"Catégorie": "Viandes", "Sous-Catégorie": "Aile de Poulet", "Unité": "Morceau"},
    {"Catégorie": "Viandes", "Sous-Catégorie": "Cuisse de poulet", "Unité": "Morceau"},
    {"Catégorie": "Viandes", "Sous-Catégorie": "Bœuf", "Unité": "Morceau"},
    {"Catégorie": "Viandes", "Sous-Catégorie": "Saucisse", "Unité": "Pièce"},
    {"Catégorie": "Viandes", "Sous-Catégorie": "Porc", "Unité": "Morceau"},
    {"Catégorie": "Viandes", "Sous-Catégorie": "Œuf", "Unité": "Alvéole"},
    {"Catégorie": "Viandes", "Sous-Catégorie": "Viandes hachées", "Unité": "Pièce"},
    # Poissons
    {"Catégorie": "Poissons & Fruits de mer", "Sous-Catégorie": "Carpe", "Unité": "Pièce"},
    {"Catégorie": "Poissons & Fruits de mer", "Sous-Catégorie": "Sole", "Unité": "Pièce"},
    {"Catégorie": "Poissons & Fruits de mer", "Sous-Catégorie": "Bar", "Unité": "Pièce"},
    {"Catégorie": "Poissons & Fruits de mer", "Sous-Catégorie": "Gambas", "Unité": "Pièce"},
    {"Catégorie": "Poissons & Fruits de mer", "Sous-Catégorie": "Crevette", "Unité": "Pièce"},
    # Légumes
    {"Catégorie": "Légumes frais", "Sous-Catégorie": "Carottes", "Unité": "Lot"},
    {"Catégorie": "Légumes frais", "Sous-Catégorie": "Pommes de terre", "Unité": "Lot"},
    {"Catégorie": "Légumes frais", "Sous-Catégorie": "Gingembre", "Unité": "Lot"},
    {"Catégorie": "Légumes frais", "Sous-Catégorie": "Choux", "Unité": "Lot"},
    {"Catégorie": "Légumes frais", "Sous-Catégorie": "Salade", "Unité": "Lot"},
    {"Catégorie": "Légumes frais", "Sous-Catégorie": "Tomates", "Unité": "Cageot"},
    {"Catégorie": "Légumes frais", "Sous-Catégorie": "Poivrons", "Unité": "Lot"},
    {"Catégorie": "Légumes frais", "Sous-Catégorie": "Aubergines", "Unité": "Lot"},
    {"Catégorie": "Légumes frais", "Sous-Catégorie": "Oignons", "Unité": "Lot"},
    {"Catégorie": "Légumes frais", "Sous-Catégorie": "Ail", "Unité": "Lot"},
    {"Catégorie": "Légumes frais", "Sous-Catégorie": "Poireaux", "Unité": "Lot"},
    {"Catégorie": "Légumes frais", "Sous-Catégorie": "Haricots verts", "Unité": "Lot"},
    {"Catégorie": "Légumes frais", "Sous-Catégorie": "Persil", "Unité": "Lot"},
    {"Catégorie": "Légumes frais", "Sous-Catégorie": "Curcuma", "Unité": "Boite"},
    {"Catégorie": "Légumes frais", "Sous-Catégorie": "Poivre", "Unité": "Lot"},
    {"Catégorie": "Légumes frais", "Sous-Catégorie": "Courgettes", "Unité": "Lot"},
    {"Catégorie": "Légumes frais", "Sous-Catégorie": "Piments", "Unité": "Lot"},
    # Fruits
    {"Catégorie": "Fruits frais", "Sous-Catégorie": "Orange", "Unité": "Pièce"},
    {"Catégorie": "Fruits frais", "Sous-Catégorie": "Papaye", "Unité": "Pièce"},
    {"Catégorie": "Fruits frais", "Sous-Catégorie": "Mangue", "Unité": "Pièce"},
    {"Catégorie": "Fruits frais", "Sous-Catégorie": "Citron", "Unité": "Pièce"},
    {"Catégorie": "Fruits frais", "Sous-Catégorie": "Goyave", "Unité": "Pièce"},
    {"Catégorie": "Fruits frais", "Sous-Catégorie": "Mandarine", "Unité": "Pièce"},
    {"Catégorie": "Fruits frais", "Sous-Catégorie": "Pastèque", "Unité": "Pièce"},
    {"Catégorie": "Fruits frais", "Sous-Catégorie": "Pamplemousse", "Unité": "Pièce"},
    {"Catégorie": "Fruits frais", "Sous-Catégorie": "Fruits pour jus", "Unité": "Pièce"},
    # Épicerie
    {"Catégorie": "Produits secs & épicerie", "Sous-Catégorie": "Riz", "Unité": "Kg"},
    {"Catégorie": "Produits secs & épicerie", "Sous-Catégorie": "Pâtes", "Unité": "gm"},
    {"Catégorie": "Produits secs & épicerie", "Sous-Catégorie": "Haricots", "Unité": "Kg"},
    {"Catégorie": "Produits secs & épicerie", "Sous-Catégorie": "Farines", "Unité": "Kg"},
    {"Catégorie": "Produits secs & épicerie", "Sous-Catégorie": "Sucres", "Unité": "Kg"},
    {"Catégorie": "Produits secs & épicerie", "Sous-Catégorie": "Sel", "Unité": "Kg"},
    {"Catégorie": "Produits secs & épicerie", "Sous-Catégorie": "Bouillons & cubes", "Unité": "Paquet"},
    {"Catégorie": "Produits secs & épicerie", "Sous-Catégorie": "Levure", "Unité": "Boite"},
    # Laitages
    {"Catégorie": "Produits laitiers & fromages", "Sous-Catégorie": "Lait & crème", "Unité": "Litre"},
    {"Catégorie": "Produits laitiers & fromages", "Sous-Catégorie": "Beurres & margarines", "Unité": "Boite"},
    {"Catégorie": "Produits laitiers & fromages", "Sous-Catégorie": "Fromages tranchés", "Unité": "Pièces"},
    {"Catégorie": "Produits laitiers & fromages", "Sous-Catégorie": "Fromages râpés", "Unité": "Pièces"},
    {"Catégorie": "Produits laitiers & fromages", "Sous-Catégorie": "Fromages en bloc", "Unité": "Pièces"},
    {"Catégorie": "Produits laitiers & fromages", "Sous-Catégorie": "Yaourts", "Unité": "Litre"},
    # Charcuterie
    {"Catégorie": "Charcuterie", "Sous-Catégorie": "Jambons", "Unité": "Tranche"},
    {"Catégorie": "Charcuterie", "Sous-Catégorie": "Bacons", "Unité": "Tranche"},
    {"Catégorie": "Charcuterie", "Sous-Catégorie": "Saucissons", "Unité": "Tranche"},
    {"Catégorie": "Charcuterie", "Sous-Catégorie": "Pepperoni", "Unité": "Tranche"},
    {"Catégorie": "Charcuterie", "Sous-Catégorie": "Charcuterie cuite", "Unité": "Tranche"},
    {"Catégorie": "Charcuterie", "Sous-Catégorie": "Charcuterie fumée", "Unité": "Tranche"},
    # Surgelés
    {"Catégorie": "Surgelés", "Sous-Catégorie": "Frites surgelées", "Unité": "Paquet"},
    {"Catégorie": "Surgelés", "Sous-Catégorie": "Légumes surgelés", "Unité": "Paquet"},
    {"Catégorie": "Surgelés", "Sous-Catégorie": "Viandes surgelées", "Unité": "Paquet"},
    {"Catégorie": "Surgelés", "Sous-Catégorie": "Poissons surgelés", "Unité": "Paquet"},
    {"Catégorie": "Surgelés", "Sous-Catégorie": "Produits panés", "Unité": "Paquet"},
    {"Catégorie": "Surgelés", "Sous-Catégorie": "Pâtes surgelées", "Unité": "Paquet"},
    # Sauces
    {"Catégorie": "Sauces & condiments", "Sous-Catégorie": "Ketchup", "Unité": "Bouteille"},
    {"Catégorie": "Sauces & condiments", "Sous-Catégorie": "Mayonnaise", "Unité": "Bouteille"},
    {"Catégorie": "Sauces & condiments", "Sous-Catégorie": "Moutarde", "Unité": "Bouteille"},
    {"Catégorie": "Sauces & condiments", "Sous-Catégorie": "Vinaigres", "Unité": "Bouteille"},
    {"Catégorie": "Sauces & condiments", "Sous-Catégorie": "Marinades", "Unité": "Bouteille"},
    {"Catégorie": "Sauces & condiments", "Sous-Catégorie": "Bouillons liquides", "Unité": "Bouteille"},
    # Boulangerie
    {"Catégorie": "Boulangerie & pains", "Sous-Catégorie": "Pains burgers", "Unité": "Nombre"},
    {"Catégorie": "Boulangerie & pains", "Sous-Catégorie": "Pains hot-dog", "Unité": "Nombre"},
    {"Catégorie": "Boulangerie & pains", "Sous-Catégorie": "Pains sandwich", "Unité": "Nombre"},
    {"Catégorie": "Boulangerie & pains", "Sous-Catégorie": "Baguettes", "Unité": "Nombre"},
    {"Catégorie": "Boulangerie & pains", "Sous-Catégorie": "Tortillas / wraps", "Unité": "Nombre"},
    {"Catégorie": "Boulangerie & pains", "Sous-Catégorie": "Pain spécial", "Unité": "Nombre"},
    # Herbes
    {"Catégorie": "Herbes & épices", "Sous-Catégorie": "Herbes fraîches", "Unité": "Bouquet"},
    {"Catégorie": "Herbes & épices", "Sous-Catégorie": "Herbes sèches", "Unité": "Bouquet"},
    {"Catégorie": "Herbes & épices", "Sous-Catégorie": "Épices moulues", "Unité": "Sachet"},
    {"Catégorie": "Herbes & épices", "Sous-Catégorie": "Épices entières", "Unité": "Sachet"},
    {"Catégorie": "Herbes & épices", "Sous-Catégorie": "Mélanges d'épices", "Unité": "Sachet"},
    # Pâtisserie
    {"Catégorie": "Produits de pâtisserie", "Sous-Catégorie": "Chocolats", "Unité": "Boite"},
    {"Catégorie": "Produits de pâtisserie", "Sous-Catégorie": "Levures", "Unité": "Sachet"},
    {"Catégorie": "Produits de pâtisserie", "Sous-Catégorie": "Arômes", "Unité": "Bouteille"},
    {"Catégorie": "Produits de pâtisserie", "Sous-Catégorie": "Sucre glace", "Unité": "Kg"},
    {"Catégorie": "Produits de pâtisserie", "Sous-Catégorie": "Cacao", "Unité": "Boite"},
    {"Catégorie": "Produits de pâtisserie", "Sous-Catégorie": "Fruits secs", "Unité": "Kg"},
    # Huiles
    {"Catégorie": "Huiles & matières grasses", "Sous-Catégorie": "Huiles végétales", "Unité": "Litre"},
    {"Catégorie": "Huiles & matières grasses", "Sous-Catégorie": "Huiles spéciales", "Unité": "Litre"},
    {"Catégorie": "Huiles & matières grasses", "Sous-Catégorie": "Beurres", "Unité": "Kg"},
    {"Catégorie": "Huiles & matières grasses", "Sous-Catégorie": "Margarines", "Unité": "Kg"},
    {"Catégorie": "Huiles & matières grasses", "Sous-Catégorie": "Graisses de friture", "Unité": "Litre"},
    # Nettoyage
    {"Catégorie": "Produits de nettoyage", "Sous-Catégorie": "Détergents", "Unité": "Bouteille"},
    {"Catégorie": "Produits de nettoyage", "Sous-Catégorie": "Désinfectants", "Unité": "Bouteille"},
    {"Catégorie": "Produits de nettoyage", "Sous-Catégorie": "Produits vaisselle", "Unité": "Bouteille"},
    {"Catégorie": "Produits de nettoyage", "Sous-Catégorie": "Produits sol", "Unité": "Bouteille"},
    {"Catégorie": "Produits de nettoyage", "Sous-Catégorie": "Accessoires (éponges, gants)", "Unité": "Lot"},
    # Emballages
    {"Catégorie": "Emballages cuisine", "Sous-Catégorie": "Barquettes", "Unité": "Lot"},
    {"Catégorie": "Emballages cuisine", "Sous-Catégorie": "Boîtes", "Unité": "Lot"},
    {"Catégorie": "Emballages cuisine", "Sous-Catégorie": "Films alimentaires", "Unité": "Rouleau"},
    {"Catégorie": "Emballages cuisine", "Sous-Catégorie": "Sacs kraft", "Unité": "Lot"},
    {"Catégorie": "Emballages cuisine", "Sous-Catégorie": "Bouteille de Gaz", "Unité": "Pièce"},
    {"Catégorie": "Emballages cuisine", "Sous-Catégorie": "Couverts jetables", "Unité": "Lot"},
    {"Catégorie": "Emballages cuisine", "Sous-Catégorie": "Charbon", "Unité": "Sac"},
]

# 3. LISTE DES CATEGORIES POUR VALIDATION
categories_list = sorted(list(set([item["Catégorie"] for item in menu_data])))
sous_categories_matières = sorted(list(set([item["Sous-Catégorie"] for item in stock_categories])))

# ==============================================================================
# FONCTIONS UTILITAIRES
# ==============================================================================

def create_header(ws, title, row=1, col=1, colspan=1):
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = fonts['title']
    cell.alignment = alignments['left']
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+colspan-1)
    return cell

def create_subtitle(ws, title, row, col=1):
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = fonts['subtitle']
    cell.alignment = alignments['left']
    return cell

def style_table_header(ws, headers, row, start_col=1):
    for i, header in enumerate(headers):
        cell = ws.cell(row=row, column=start_col+i, value=header)
        cell.font = fonts['header']
        cell.fill = PatternFill(start_color=colors['primary'], end_color=colors['primary'], fill_type="solid")
        cell.alignment = alignments['center']
        cell.border = borders['thin']

def create_named_range(wb, name, sheet_name, ref):
    try:
        wb.defined_names.add(openpyxl.workbook.defined_name.DefinedName(name, attr_text=f"'{sheet_name}'!{ref}"))
    except Exception as e:
        print(f"Erreur création nom {name}: {e}")

# ==============================================================================
# FEUILLE 1: ACCUEIL
# ==============================================================================
ws_accueil = wb.create_sheet("ACCUEIL")
ws_accueil.sheet_view.showGridLines = False

create_header(ws_accueil, "🍽️ GESTION DU RESTAURANT", 1, 1, 5)
ws_accueil.cell(2, 1, value="Application Professionnelle de Gestion - Version 3.0").font = fonts['normal']

# Menu
create_subtitle(ws_accueil, "📂 MODULES PRINCIPAUX", 4, 1)
modules = [
    ("📊 Dashboard", "DASHBOARD"),
    ("🍔 Carte & Produits", "PRODUITS"),
    ("📦 Matières Premières & Stocks", "MATIERES_PREMIERES"),
    ("👨‍🍳 Recettes & Coûts", "RECETTES"),
    ("💰 Ventes Journalières", "VENTES"),
    ("🚚 Fournisseurs & Achats", "FOURNISSEURS"),
    ("📝 Inventaires", "INVENTAIRES"),
    ("📈 Tableaux de Bord", "TABLEAUX_BORD"),
    ("⚙️ Catégories", "CATEGORIES"),
    ("❓ Aide", "AIDE")
]

for i, (txt, sheet) in enumerate(modules):
    row = 6 + i
    cell = ws_accueil.cell(row=row, column=2, value=f"  {txt}  ")
    cell.font = fonts['link']
    cell.fill = PatternFill(start_color=colors['light'], end_color=colors['light'], fill_type="solid")
    cell.border = borders['thin']
    cell.alignment = alignments['left']
    # Note: Les liens hypertextes réels nécessitent VBA ou une manipulation XML complexe ici, 
    # mais dans Excel, on peut les ajouter manuellement ou via une macro simple si besoin.
    # Pour l'instant, c'est visuel.

# Alertes
create_subtitle(ws_accueil, "🔔 ALERTES EN TEMPS RÉEL", 4, 8)
ws_accueil.cell(6, 8, value="⚠️ Ruptures de stock").font = fonts['bold']
ws_accueil.cell(7, 8, value="• Aucun produit en rupture actuellement").font = fonts['normal']
ws_accueil.cell(9, 8, value="📉 Ventes du jour").font = fonts['bold']
ws_accueil.cell(10, 8, value="• En attente de saisie").font = fonts['normal']

ws_accueil.column_dimensions['A'].width = 5
ws_accueil.column_dimensions['B'].width = 40
ws_accueil.column_dimensions['H'].width = 40

# ==============================================================================
# FEUILLE 2: CATEGORIES
# ==============================================================================
ws_cat = wb.create_sheet("CATEGORIES")
headers_cat = ["ID", "Nom de la Catégorie", "Description", "Nombre de Produits"]
style_table_header(ws_cat, headers_cat, 3)

# Remplissage automatique
for i, cat in enumerate(categories_list, start=4):
    ws_cat.cell(i, 1, value=i-3)
    ws_cat.cell(i, 2, value=cat)
    ws_cat.cell(i, 3, value=f"Gestion de {cat}")
    # Formule pour compter les produits (à ajuster selon la structure finale)
    ws_cat.cell(i, 4, value=f'=NB.SI(PRODUITS!C:C; "{cat}")') 

ws_cat.column_dimensions['A'].width = 5
ws_cat.column_dimensions['B'].width = 25
ws_cat.column_dimensions['C'].width = 40
ws_cat.column_dimensions['D'].width = 15

# ==============================================================================
# FEUILLE 3: PRODUITS (CARTE)
# ==============================================================================
ws_prod = wb.create_sheet("PRODUITS")
headers_prod = ["ID", "Nom du Plat", "Catégorie", "Variante/Détail", "Prix Vente (FCFA)", "Observation", "Statut", "Date Maj"]
style_table_header(ws_prod, headers_prod, 3)

for i, item in enumerate(menu_data, start=4):
    ws_prod.cell(i, 1, value=i-3)
    ws_prod.cell(i, 2, value=item["Nom"])
    ws_prod.cell(i, 3, value=item["Catégorie"])
    ws_prod.cell(i, 4, value=item.get("Variante", ""))
    ws_prod.cell(i, 5, value=item["Prix"])
    ws_prod.cell(i, 6, value=item["Observation"])
    ws_prod.cell(i, 7, value="Actif")
    ws_prod.cell(i, 8, value=datetime.date.today())

# Formatage
for row in ws_prod.iter_rows(min_row=4, max_row=len(menu_data)+3, min_col=5, max_col=5):
    for cell in row:
        cell.number_format = '#,##0'

ws_prod.column_dimensions['A'].width = 5
ws_prod.column_dimensions['B'].width = 30
ws_prod.column_dimensions['C'].width = 20
ws_prod.column_dimensions['D'].width = 25
ws_prod.column_dimensions['E'].width = 15
ws_prod.column_dimensions['F'].width = 30

# Validation de données pour la catégorie
from openpyxl.worksheet.datavalidation import DataValidation
dv_cat = DataValidation(type="list", formula1=f"CATEGORIES!$B$4:$B${len(categories_list)+3}", allow_blank=True)
dv_cat.error = "Veuillez sélectionner une catégorie existante."
ws_prod.add_data_validation(dv_cat)
dv_cat.add(ws_prod.cell(row=4, column=3)) # Appliqué à partir de la première ligne de donnée

# ==============================================================================
# FEUILLE 4: MATIERES PREMIERES (STOCK DETAILLÉ)
# ==============================================================================
ws_stock = wb.create_sheet("MATIERES_PREMIERES")
headers_stock = ["ID", "Catégorie", "Sous-Catégorie", "Unité", "Stock Initial", "Entrées", "Sorties", "Stock Final", "Coût Unitaire", "Valeur Totale"]
style_table_header(ws_stock, headers_stock, 3)

for i, item in enumerate(stock_categories, start=4):
    ws_stock.cell(i, 1, value=i-3)
    ws_stock.cell(i, 2, value=item["Catégorie"])
    ws_stock.cell(i, 3, value=item["Sous-Catégorie"])
    ws_stock.cell(i, 4, value=item["Unité"])
    ws_stock.cell(i, 5, value=0) # Stock initial à remplir
    ws_stock.cell(i, 6, value=0)
    ws_stock.cell(i, 7, value=0)
    # Formule Stock Final
    ws_stock.cell(i, 8, value=f"=E{i}+F{i}-G{i}")
    ws_stock.cell(i, 9, value=0) # Coût à remplir
    # Formule Valeur Totale
    ws_stock.cell(i, 10, value=f"=H{i}*I{i}")

# Formatage
for row in ws_stock.iter_rows(min_row=4, max_row=len(stock_categories)+3, min_col=5, max_col=10):
    for cell in row:
        if cell.column != 4: # Pas sur la colonne Unité
            cell.number_format = '#,##0'

ws_stock.column_dimensions['A'].width = 5
ws_stock.column_dimensions['B'].width = 25
ws_stock.column_dimensions['C'].width = 25
ws_stock.column_dimensions['D'].width = 10
ws_stock.column_dimensions['E'].width = 12
ws_stock.column_dimensions['F'].width = 12
ws_stock.column_dimensions['G'].width = 12
ws_stock.column_dimensions['H'].width = 12
ws_stock.column_dimensions['I'].width = 15
ws_stock.column_dimensions['J'].width = 15

# Mise en forme conditionnelle pour stock faible (< 5)
rule_low = CellIsRule(operator='lessThan', formula=['5'], fill=PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid'))
ws_stock.conditional_formatting.add(f'H4:H{len(stock_categories)+3}', rule_low)

# ==============================================================================
# FEUILLE 5: RECETTES
# ==============================================================================
ws_rec = wb.create_sheet("RECETTES")
headers_rec = ["ID", "Nom de la Recette", "Catégorie", "Coût Matière Total", "Prix Vente Conseillé", "Marge Brute", "% Marge"]
style_table_header(ws_rec, headers_rec, 3)

# Exemple de formules liées aux ingrédients (à compléter dans l'onglet Ingrédients)
# Pour l'instant, on met des placeholders
for i, item in enumerate(menu_data[:10], start=4): # Juste les 10 premiers pour exemple
    ws_rec.cell(i, 1, value=i-3)
    ws_rec.cell(i, 2, value=item["Nom"])
    ws_rec.cell(i, 3, value=item["Catégorie"])
    ws_rec.cell(i, 4, value=0) # À calculer via SOMMEPROD avec l'onglet Ingrédients
    ws_rec.cell(i, 5, value=item["Prix"])
    ws_rec.cell(i, 6, value=f"=E{i}-D{i}")
    ws_rec.cell(i, 7, value=f"=SI(E{i}>0; F{i}/E{i}; 0)")

ws_rec.column_dimensions['A'].width = 5
ws_rec.column_dimensions['B'].width = 30
ws_rec.column_dimensions['C'].width = 20
ws_rec.column_dimensions['D'].width = 15
ws_rec.column_dimensions['E'].width = 15
ws_rec.column_dimensions['F'].width = 15
ws_rec.column_dimensions['G'].width = 10

# ==============================================================================
# FEUILLE 6: INGREDIENTS (Détail des recettes)
# ==============================================================================
ws_ing = wb.create_sheet("INGREDIENTS")
headers_ing = ["Recette", "Matière Première", "Quantité Requise", "Unité", "Coût Unitaire MP", "Coût Ligne"]
style_table_header(ws_ing, headers_ing, 3)

ws_ing.column_dimensions['A'].width = 25
ws_ing.column_dimensions['B'].width = 30
ws_ing.column_dimensions['C'].width = 15
ws_ing.column_dimensions['D'].width = 10
ws_ing.column_dimensions['E'].width = 15
ws_ing.column_dimensions['F'].width = 15

# ==============================================================================
# FEUILLE 7: VENTES (Adaptée FICHE VENTES)
# ==============================================================================
ws_vente = wb.create_sheet("VENTES")
create_header(ws_vente, "📅 FICHE VENTES JOURNALIÈRES", 1, 1, 6)
ws_vente.cell(2, 1, value="Date :").font = fonts['bold']
ws_vente.cell(2, 2, value=datetime.date.today()).number_format = 'dd/mm/yyyy'
ws_vente.cell(2, 4, value="Heure Début :").font = fonts['bold']
ws_vente.cell(2, 6, value="Heure Fin :").font = fonts['bold']

headers_vente = ["ID", "Plat Vendu", "Catégorie", "Quantité", "Prix Unitaire", "Total FCFA", "Observation"]
style_table_header(ws_vente, headers_vente, 4)

# Lignes vides pour saisie
for i in range(5, 25):
    ws_vente.cell(i, 1, value=i-4)
    # Liste déroulante pour le plat
    # Formule pour récupérer le prix automatiquement (si le nom correspond)
    # =SI(B5<>""; INDEX(PRODUITS!E:E; EQUIV(B5; PRODUITS!B:B; 0)); "")
    ws_vente.cell(i, 5, value=f'=SI($B{i}<>""; INDEX(PRODUITS!$E:$E; EQUIV($B{i}; PRODUITS!$B:$B; 0)); "")')
    # Total
    ws_vente.cell(i, 6, value=f'=IF(AND($D{i}>0; $E{i}>0); $D{i}*$E{i}; "")')
    
    # Format
    ws_vente.cell(i, 5).number_format = '#,##0'
    ws_vente.cell(i, 6).number_format = '#,##0'

# Totaux en bas
row_total = 26
ws_vente.cell(row_total, 4, value="TOTAL VENTES :").font = fonts['bold']
ws_vente.cell(row_total, 6, value=f'=SUM(F5:F25)').font = fonts['bold']
ws_vente.cell(row_total, 6).number_format = '#,##0 FCFA'

ws_vente.column_dimensions['A'].width = 5
ws_vente.column_dimensions['B'].width = 35
ws_vente.column_dimensions['C'].width = 20
ws_vente.column_dimensions['D'].width = 10
ws_vente.column_dimensions['E'].width = 15
ws_vente.column_dimensions['F'].width = 15
ws_vente.column_dimensions['G'].width = 30

# Validation pour les plats
dv_plat = DataValidation(type="list", formula1="PRODUITS!$B$4:$B$100", allow_blank=True)
ws_vente.add_data_validation(dv_plat)
dv_plat.add(ws_vente.cell(row=5, column=2))

# Signature
ws_vente.cell(28, 1, value="Signature pour Validation :").font = fonts['bold']
ws_vente.merge_cells(start_row=28, start_column=2, end_row=28, end_column=4)
ws_vente.cell(28, 2, value="__________________________________")

# ==============================================================================
# FEUILLE 8: FOURNISSEURS
# ==============================================================================
ws_fourn = wb.create_sheet("FOURNISSEURS")
headers_fourn = ["ID", "Nom Fournisseur", "Contact", "Téléphone", "Email", "Adresse", "Spécialité"]
style_table_header(ws_fourn, headers_fourn, 3)
ws_fourn.column_dimensions['A'].width = 5
ws_fourn.column_dimensions['B'].width = 25
ws_fourn.column_dimensions['C'].width = 25
ws_fourn.column_dimensions['D'].width = 15
ws_fourn.column_dimensions['E'].width = 25
ws_fourn.column_dimensions['F'].width = 30
ws_fourn.column_dimensions['G'].width = 20

# ==============================================================================
# FEUILLE 9: INVENTAIRES
# ==============================================================================
ws_inv = wb.create_sheet("INVENTAIRES")
create_header(ws_inv, "📝 FICHE INVENTAIRE PÉRIODIQUE", 1, 1, 6)
ws_inv.cell(2, 1, value="Date :").font = fonts['bold']
ws_inv.cell(2, 2, value=datetime.date.today()).number_format = 'dd/mm/yyyy'

headers_inv = ["Matière", "Catégorie", "Unité", "Stock Théorique", "Stock Réel", "Écart", "Type Écart", "Coût Unitaire", "Valeur Écart"]
style_table_header(ws_inv, headers_inv, 4)

# Exemple de lignes liées aux matières premières
for i in range(5, 15):
    ws_inv.cell(i, 2, value="À sélectionner")
    ws_inv.cell(i, 4, value=0) # Théorique (lié au stock)
    ws_inv.cell(i, 5, value=0) # Réel (saisie)
    ws_inv.cell(i, 6, value=f"=E{i}-D{i}")
    ws_inv.cell(i, 7, value=f'=SI(F{i}=0; "OK"; SI(F{i}>0; "Excédent"; "Perte"))')
    ws_inv.cell(i, 8, value=0)
    ws_inv.cell(i, 9, value=f"=F{i}*H{i}")

ws_inv.column_dimensions['A'].width = 25
ws_inv.column_dimensions['B'].width = 20
ws_inv.column_dimensions['C'].width = 10
ws_inv.column_dimensions['D'].width = 15
ws_inv.column_dimensions['E'].width = 15
ws_inv.column_dimensions['F'].width = 10
ws_inv.column_dimensions['G'].width = 12
ws_inv.column_dimensions['H'].width = 15
ws_inv.column_dimensions['I'].width = 15

# ==============================================================================
# FEUILLE 10: DASHBOARD
# ==============================================================================
ws_dash = wb.create_sheet("DASHBOARD")
create_header(ws_dash, "📊 TABLEAU DE BORD", 1, 1, 4)

# KPIs
kpis = [
    ("Ventes du Jour", "VENTES!F26", colors['accent']),
    ("Marge Estimée", "0 FCFA", colors['secondary']), # Formule complexe à faire manuellement
    ("Nombre de Ventes", "0", colors['warning']),
    ("Alertes Stock", "Voir Onglet Stocks", colors['danger'])
]

for i, (label, val, color) in enumerate(kpis):
    row = 4 + (i // 2) * 4
    col = 2 + (i % 2) * 4
    
    ws_dash.cell(row, col, value=label).font = fonts['subtitle']
    cell_val = ws_dash.cell(row+1, col, value=val)
    cell_val.font = Font(size=16, bold=True, color=color)
    cell_val.alignment = alignments['center']
    
    # Cadre
    for r in range(row, row+3):
        for c in range(col, col+3):
            ws_dash.cell(r, c).border = borders['thin']

ws_dash.column_dimensions['B'].width = 20
ws_dash.column_dimensions['F'].width = 20

# ==============================================================================
# FEUILLE 11: TABLEAUX_BORD
# ==============================================================================
ws_tb = wb.create_sheet("TABLEAUX_BORD")
create_header(ws_tb, "📈 ANALYTIQUES AVANCÉS", 1, 1, 4)
ws_tb.cell(3, 1, value="Les tableaux croisés dynamiques seront générés ici une fois les données saisies.").font = fonts['normal']
ws_tb.cell(4, 1, value="Allez dans Insertion > Tableau Croisé Dynamique pour analyser les ventes et stocks.").font = fonts['normal']

# ==============================================================================
# FEUILLE 12: AIDE
# ==============================================================================
ws_aide = wb.create_sheet("AIDE")
create_header(ws_aide, "❓ GUIDE D'UTILISATION RAPIDE", 1, 1, 4)

help_text = [
    "1. COMMENCER :",
    "- Allez dans l'onglet 'CATEGORIES' pour vérifier vos catégories.",
    "- Allez dans 'PRODUITS' : Vos plats sont déjà pré-enregistrés ! Vérifiez les prix.",
    "- Allez dans 'MATIERES_PREMIERES' : Saisissez vos stocks initiaux et coûts unitaires.",
    "",
    "2. QUOTIDIEN :",
    "- Enregistrement des achats : Mettez à jour les 'Entrées' dans 'MATIERES_PREMIERES'.",
    "- Enregistrement des ventes : Utilisez l'onglet 'VENTES'. Sélectionnez le plat, la quantité se calcule seule.",
    "",
    "3. SUIVI :",
    "- Consultez le 'DASHBOARD' pour les totaux du jour.",
    "- Faites l'inventaire dans 'INVENTAIRES' en fin de semaine/mois.",
    "",
    "4. ASTUCES :",
    "- Les listes déroulantes vous aident à ne pas faire d'erreur de frappe.",
    "- Ne modifiez pas les colonnes avec des formules (texte en gris ou calcul automatique)."
]

for i, line in enumerate(help_text, start=3):
    cell = ws_aide.cell(i, 1, value=line)
    if line.endswith(":"):
        cell.font = fonts['bold']
    else:
        cell.font = fonts['normal']

ws_aide.column_dimensions['A'].width = 80

# ==============================================================================
# SAUVEGARDE
# ==============================================================================
wb.save(filename)
print(f"✅ Application Excel générée avec succès : {filename}")
print(f"   - {len(menu_data)} plats intégrés dans l'onglet PRODUITS")
print(f"   - {len(stock_categories)} matières premières intégrées dans l'onglet MATIERES_PREMIERES")
print(f"   - Feuille VENTES adaptée à votre format journalier")
print(f"   - Listes déroulantes configurées")
