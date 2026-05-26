#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Création de l'application Excel de gestion de restaurant
Sans VBA - Utilisation de formules avancées, tableaux structurés et mise en forme conditionnelle
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
import json

# Couleurs professionnelles
COLORS = {
    'primary': '1F4E79',      # Bleu foncé
    'secondary': '2E86AB',    # Bleu moyen
    'accent': 'A23B72',       # Rose/Violet
    'success': '28A745',      # Vert
    'warning': 'FFC107',      # Jaune
    'danger': 'DC3545',       # Rouge
    'light': 'F8F9FA',        # Gris clair
    'dark': '343A40',         # Gris foncé
    'white': 'FFFFFF'
}

# Styles
FONTS = {
    'title': Font(name='Calibri', size=16, bold=True, color=Color(rgb='FFFFFF')),
    'subtitle': Font(name='Calibri', size=12, bold=True, color=Color(rgb='1F4E79')),
    'header': Font(name='Calibri', size=11, bold=True, color=Color(rgb='FFFFFF')),
    'normal': Font(name='Calibri', size=10),
    'alert': Font(name='Calibri', size=10, bold=True, color=Color(rgb='DC3545'))
}

ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_header_row(ws, row, headers, start_col=1):
    """Crée une ligne d'en-tête stylisée"""
    fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    
    for i, header in enumerate(headers):
        col = start_col + i
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = FONTS['header']
        cell.fill = fill
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    
    return len(headers)

def create_table(ws, name, start_row, start_col, num_rows, num_cols, headers):
    """Crée un tableau structuré Excel"""
    # Créer les en-têtes
    create_header_row(ws, start_row, headers, start_col)
    
    # Définir la plage du tableau
    end_row = start_row + num_rows
    end_col = start_col + num_cols - 1
    
    table_ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    
    # Créer le tableau
    table = Table(displayName=name, ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    
    return table

def setup_sheet(ws, sheet_name, zoom=100):
    """Configure les paramètres de base d'une feuille"""
    ws.sheet_view.zoomScale = zoom
    ws.sheet_view.zoomScaleNormal = zoom
    ws.freeze_panes = 'A2'
    
    # Largeurs de colonnes par défaut
    ws.column_dimensions['A'].width = 5
    for col in range(2, 20):
        ws.column_dimensions[get_column_letter(col)].width = 15

def create_home_sheet(wb):
    """Crée la page d'accueil"""
    ws = wb.create_sheet('ACCUEIL')
    ws.sheet_view.showGridLines = False
    
    # Titre principal
    ws.merge_cells('A1:H1')
    title = ws.cell(row=1, column=1, value='🍽️ GESTION DE RESTAURANT PROFESSIONNEL')
    title.font = Font(name='Calibri', size=20, bold=True, color=Color(rgb='FFFFFF'))
    title.fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    title.alignment = ALIGN_CENTER
    
    # Sous-titre
    ws.merge_cells('A2:H2')
    subtitle = ws.cell(row=2, column=1, value='Application Excel Complète - Sans VBA')
    subtitle.font = Font(name='Calibri', size=12, italic=True)
    subtitle.alignment = ALIGN_CENTER
    
    # Section Navigation
    ws.merge_cells('A4:H4')
    nav_title = ws.cell(row=4, column=1, value='📋 NAVIGATION PRINCIPALE')
    nav_title.font = FONTS['subtitle']
    nav_title.alignment = ALIGN_CENTER
    
    modules = [
        ('A6', 'B6', '📊 DASHBOARD', 'DASHBOARD', 'Tableau de bord avec KPI'),
        ('A8', 'B8', '📦 PRODUITS', 'PRODUITS', 'Gestion des produits et matières'),
        ('A10', 'B10', '🏷️ CATÉGORIES', 'CATEGORIES', 'Catégories et sous-catégories'),
        ('A12', 'B12', '📈 STOCKS', 'STOCKS', 'Suivi des mouvements de stock'),
        ('C6', 'D6', '📖 RECETTES', 'RECETTES', 'Gestion des recettes et coûts'),
        ('C8', 'D8', '🛒 INGREDIENTS', 'INGREDIENTS', 'Liste des ingrédients'),
        ('C10', 'D10', '💰 VENTES', 'VENTES', 'Enregistrement des ventes'),
        ('C12', 'D12', '🚚 FOURNISSEURS', 'FOURNISSEURS', 'Gestion fournisseurs et achats'),
        ('E6', 'F6', '📝 INVENTAIRES', 'INVENTAIRES', 'Contrôles périodiques'),
        ('E8', 'F8', '📊 TABLEAUX BORD', 'TABLEAUX_BORD', 'Analyses détaillées'),
        ('E10', 'F10', '⚙️ PARAMÈTRES', 'PARAMETRES', 'Configuration système'),
        ('E12', 'F12', '❓ AIDE', 'AIDE', 'Guide d\'utilisation')
    ]
    
    fill_blue = PatternFill(start_color=COLORS['secondary'], end_color=COLORS['secondary'], fill_type='solid')
    fill_accent = PatternFill(start_color=COLORS['accent'], end_color=COLORS['accent'], fill_type='solid')
    
    for col1, col2, text, link, desc in modules:
        # Bouton
        ws.merge_cells(f'{col1}:{col2}')
        btn = ws.cell(row=int(col1[1:]), column=ord(col1[0])-64, value=text)
        btn.font = Font(name='Calibri', size=11, bold=True, color=Color(rgb='FFFFFF'))
        btn.fill = fill_blue if 'DASHBOARD' not in text else fill_accent
        btn.alignment = ALIGN_CENTER
        btn.border = THIN_BORDER
        
        # Description
        desc_cell = ws.cell(row=int(col1[1:])+1, column=ord(col1[0])-64, value=desc)
        desc_cell.font = Font(name='Calibri', size=9, italic=True)
        desc_cell.alignment = ALIGN_CENTER
    
    # Section Alertes
    ws.merge_cells('A16:H16')
    alert_title = ws.cell(row=16, column=1, value='🔔 ALERTES IMPORTANTES')
    alert_title.font = FONTS['subtitle']
    alert_title.alignment = ALIGN_CENTER
    
    # Tableau des alertes (formules)
    alert_headers = ['Type', 'Message', 'Priorité', 'Action Requise']
    create_header_row(ws, 17, alert_headers, 1)
    
    # Lignes d'alertes avec formules
    alerts_data = [
        ['Stock', '=IF(COUNTIF(PRODUITS!H:H,"⚠️ RUPTURE")>0, COUNTIF(PRODUITS!H:H,"⚠️ RUPTURE")&" ruptures de stock", "Aucune")', 
         '=IF(COUNTIF(PRODUITS!H:H,"⚠️ RUPTURE")>0,"HAUTE","NORMALE")', 'Vérifier stocks'],
        ['Commandes', '=IF(COUNTIF(FOURNISSEURS!G:G,"En attente")>0, COUNTIF(FOURNISSEURS!G:G,"En attente")&" commandes", "Aucune")',
         '=IF(COUNTIF(FOURNISSEURS!G:G,"En attente")>0,"MOYENNE","NORMALE")', 'Suivre commandes'],
        ['Inventaire', '=IF(TODAY()-MAX(INVENTAIRES!A:A)>7,"Plus de 7 jours","À jour")',
         '=IF(TODAY()-MAX(INVENTAIRES!A:A)>7,"HAUTE","NORMALE")', 'Planifier inventaire']
    ]
    
    for i, alert in enumerate(alerts_data, 18):
        for j, val in enumerate(alert, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = FONTS['normal']
            cell.alignment = ALIGN_CENTER if j > 1 else ALIGN_LEFT
            cell.border = THIN_BORDER
    
    # Pied de page
    ws.merge_cells('A22:H22')
    footer = ws.cell(row=22, column=1, value='© 2024 - Application de Gestion de Restaurant | Version 1.0')
    footer.font = Font(name='Calibri', size=9, italic=True)
    footer.alignment = ALIGN_CENTER
    
    # Ajuster largeurs
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 3
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 15
    
    return ws

def create_dashboard_sheet(wb):
    """Crée le tableau de bord principal"""
    ws = wb.create_sheet('DASHBOARD')
    
    # Titre
    ws.merge_cells('A1:L1')
    title = ws.cell(row=1, column=1, value='📊 TABLEAU DE BORD PRINCIPAL')
    title.font = Font(name='Calibri', size=18, bold=True, color=Color(rgb='FFFFFF'))
    title.fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    title.alignment = ALIGN_CENTER
    
    # KPI Section
    kpi_data = [
        ('B3', 'Ventes Totales', '=SUM(VENTES!F:F)', '€', COLORS['success']),
        ('D3', 'Marge Brute', '=SUM(VENTES!G:G)', '€', COLORS['accent']),
        ('F3', 'Taux Marge', '=IFERROR(D3/B3*100,0)', '%', COLORS['secondary']),
        ('H3', 'Nb Ventes', '=COUNTA(VENTES!A:A)-1', '', COLORS['warning']),
        ('J3', 'Panier Moyen', '=IFERROR(B3/H3,0)', '€', COLORS['primary']),
        ('B6', 'Valeur Stock', '=SUMPRODUCT(PRODUITS!F:F,PRODUITS!G:G)', '€', COLORS['info'] if hasattr(Color, 'info') else '17A2B8'),
        ('D6', 'Produits Critiques', '=COUNTIF(PRODUITS!H:H,"⚠️ CRITIQUE")+COUNTIF(PRODUITS!H:H,"⚠️ RUPTURE")', '', COLORS['danger']),
        ('F6', 'Recettes Actives', '=COUNTA(RECETTES!A:A)-1', '', COLORS['success']),
        ('H6', 'Fournisseurs', '=COUNTA(FOURNISSEURS!A:A)-1', '', COLORS['secondary']),
        ('J6', 'Dernier Inventaire', '=MAX(INVENTAIRES!A:A)', 'date', COLORS['warning'])
    ]
    
    for cell_ref, label, formula, unit, color in kpi_data:
        col = ord(cell_ref[0]) - 64
        row = int(cell_ref[1:])
        
        # Label
        label_cell = ws.cell(row=row, column=col, value=label)
        label_cell.font = Font(name='Calibri', size=9, bold=True)
        label_cell.alignment = ALIGN_CENTER
        
        # Valeur
        val_cell = ws.cell(row=row+1, column=col, value=formula)
        val_cell.font = Font(name='Calibri', size=14, bold=True, color=Color(rgb=color))
        val_cell.alignment = ALIGN_CENTER
        val_cell.number_format = '#,##0.00' if unit == '€' else '0.00%' if unit == '%' else 'dd/mm/yyyy' if unit == 'date' else '#,##0'
        
        # Bordure
        for c in range(col, col+1):
            for r in range(row, row+3):
                ws.cell(row=r, column=c).border = THIN_BORDER
    
    # Section Graphiques (placeholders)
    ws.merge_cells('A10:L10')
    chart_title = ws.cell(row=10, column=1, value='📈 GRAPHIQUES ET ANALYSES')
    chart_title.font = FONTS['subtitle']
    chart_title.alignment = ALIGN_CENTER
    
    instructions = ws.cell(row=11, column=1, value='''Pour créer les graphiques :
1. Sélectionnez les données dans les feuilles correspondantes
2. Insertion → Graphiques recommandés
3. Types suggérés : Courbes pour l'évolution, Camembert pour les répartitions, Histogrammes pour les comparaisons''')
    instructions.font = Font(name='Calibri', size=10, italic=True)
    instructions.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    ws.row_dimensions[11].height = 80
    
    return ws

def create_products_sheet(wb):
    """Crée la feuille de gestion des produits"""
    ws = wb.create_sheet('PRODUITS')
    
    # Titre
    ws.merge_cells('A1:K1')
    title = ws.cell(row=1, column=1, value='📦 GESTION DES PRODUITS ET MATIÈRES PREMIÈRES')
    title.font = Font(name='Calibri', size=16, bold=True, color=Color(rgb='FFFFFF'))
    title.fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    title.alignment = ALIGN_CENTER
    
    # En-têtes
    headers = ['ID', 'Nom Produit', 'Description', 'Catégorie', 'Sous-Catégorie', 
               'Prix Achat (€)', 'Stock Actuel', 'Statut Stock', 'Stock Min', 'Stock Sécurité', 'Unité']
    
    create_header_row(ws, 2, headers, 1)
    
    # Données exemple
    sample_data = [
        ['P001', 'Tomates', 'Tomates fraîches', 'Légumes', 'Frais', 2.50, 50, '', 20, 10, 'kg'],
        ['P002', 'Boeuf Haché', 'Qualité premium', 'Viandes', 'Frais', 12.00, 15, '', 10, 5, 'kg'],
        ['P003', 'Riz Basmati', 'Riz parfumé', 'Épicerie', 'Sec', 3.20, 100, '', 30, 15, 'kg'],
        ['P004', 'Huile Olive', 'Extra vierge', 'Épicerie', 'Sec', 8.50, 20, '', 10, 5, 'L'],
        ['P005', 'Saumon', 'Frais Atlantic', 'Poissons', 'Frais', 18.00, 8, '', 10, 5, 'kg']
    ]
    
    for i, data in enumerate(sample_data, 3):
        for j, val in enumerate(data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = FONTS['normal']
            cell.alignment = ALIGN_CENTER if j in [1, 6, 7, 9, 10] else ALIGN_LEFT
            cell.border = THIN_BORDER
            
            # Formule statut stock (colonne H)
            if j == 8:
                cell.value = '=IF(G2=0,"⚠️ RUPTURE",IF(G2<D2,"⚠️ CRITIQUE",IF(G2<E2,"⚡ ATTENTION","✓ OK")))'
    
    # Ajuster largeurs
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 8
    
    # Créer tableau
    create_table(ws, 'TableauProduits', 2, 1, 100, 11, headers)
    
    return ws

def create_categories_sheet(wb):
    """Crée la feuille de gestion des catégories"""
    ws = wb.create_sheet('CATEGORIES')
    
    # Titre
    ws.merge_cells('A1:E1')
    title = ws.cell(row=1, column=1, value='🏷️ GESTION DES CATÉGORIES ET SOUS-CATÉGORIES')
    title.font = Font(name='Calibri', size=16, bold=True, color=Color(rgb='FFFFFF'))
    title.fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    title.alignment = ALIGN_CENTER
    
    # En-têtes
    headers = ['ID Catégorie', 'Nom Catégorie', 'ID Sous-Catégorie', 'Nom Sous-Catégorie', 'Description']
    create_header_row(ws, 2, headers, 1)
    
    # Données exemple
    sample_data = [
        ['CAT01', 'Légumes', 'SC01', 'Frais', 'Légumes frais de saison'],
        ['CAT01', 'Légumes', 'SC02', 'Surgelés', 'Légumes surgelés'],
        ['CAT02', 'Viandes', 'SC03', 'Frais', 'Viandes fraîches'],
        ['CAT02', 'Viandes', 'SC04', 'Charcuterie', 'Produits de charcuterie'],
        ['CAT03', 'Poissons', 'SC05', 'Frais', 'Poissons et fruits de mer frais'],
        ['CAT04', 'Épicerie', 'SC06', 'Sec', 'Produits secs'],
        ['CAT04', 'Épicerie', 'SC07', 'Conserves', 'Produits en conserve'],
        ['CAT05', 'Boissons', 'SC08', 'Sans alcool', 'Boissons non alcoolisées'],
        ['CAT05', 'Boissons', 'SC09', 'Alcool', 'Vins et spiritueux']
    ]
    
    for i, data in enumerate(sample_data, 3):
        for j, val in enumerate(data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = FONTS['normal']
            cell.alignment = ALIGN_CENTER if j <= 3 else ALIGN_LEFT
            cell.border = THIN_BORDER
    
    # Ajuster largeurs
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30
    
    create_table(ws, 'TableauCategories', 2, 1, 50, 5, headers)
    
    return ws

def create_stocks_sheet(wb):
    """Crée la feuille de gestion des stocks"""
    ws = wb.create_sheet('STOCKS')
    
    # Titre
    ws.merge_cells('A1:I1')
    title = ws.cell(row=1, column=1, value='📈 GESTION DES MOUVEMENTS DE STOCK')
    title.font = Font(name='Calibri', size=16, bold=True, color=Color(rgb='FFFFFF'))
    title.fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    title.alignment = ALIGN_CENTER
    
    # En-têtes
    headers = ['Date', 'ID Mouvement', 'Type', 'ID Produit', 'Nom Produit', 'Quantité', 
               'Prix Unitaire', 'Total', 'Fournisseur/Client', 'Commentaire']
    create_header_row(ws, 2, headers, 1)
    
    # Données exemple
    sample_data = [
        ['01/01/2024', 'MVT001', 'Entrée', 'P001', 'Tomates', 100, 2.50, 250.00, 'Fournisseur A', 'Livraison hebdo'],
        ['02/01/2024', 'MVT002', 'Sortie', 'P001', 'Tomates', -30, 2.50, -75.00, 'Cuisine', 'Consommation'],
        ['03/01/2024', 'MVT003', 'Entrée', 'P002', 'Boeuf Haché', 50, 12.00, 600.00, 'Fournisseur B', 'Livraison fraîche'],
        ['04/01/2024', 'MVT004', 'Sortie', 'P002', 'Boeuf Haché', -20, 12.00, -240.00, 'Cuisine', 'Consommation'],
        ['05/01/2024', 'MVT005', 'Ajustement', 'P003', 'Riz Basmati', -5, 3.20, -16.00, 'Inventaire', 'Perte constatée']
    ]
    
    for i, data in enumerate(sample_data, 3):
        for j, val in enumerate(data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = FONTS['normal']
            cell.alignment = ALIGN_CENTER if j in [1, 2, 3, 4, 6, 7, 8] else ALIGN_LEFT
            cell.border = THIN_BORDER
            if j == 8:
                cell.number_format = '#,##0.00 €'
    
    # Formule pour le total (colonne H)
    ws.cell(row=3, column=8, value='=F3*G3')
    
    # Ajuster largeurs
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 25
    
    create_table(ws, 'TableauStocks', 2, 1, 200, 10, headers)
    
    return ws

def create_recettes_sheet(wb):
    """Crée la feuille de gestion des recettes"""
    ws = wb.create_sheet('RECETTES')
    
    # Titre
    ws.merge_cells('A1:J1')
    title = ws.cell(row=1, column=1, value='📖 GESTION DES RECETTES ET COÛTS MATIÈRES')
    title.font = Font(name='Calibri', size=16, bold=True, color=Color(rgb='FFFFFF'))
    title.fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    title.alignment = ALIGN_CENTER
    
    # En-têtes
    headers = ['ID Recette', 'Nom Plat', 'Catégorie', 'Prix Vente', 'Coût Matière', 
               'Marge Brute', 'Taux Marge', 'Portions', 'Temps Prep', 'Statut']
    create_header_row(ws, 2, headers, 1)
    
    # Données exemple
    sample_data = [
        ['R001', 'Burger Maison', 'Plats Principaux', 15.00, '', '', '', 1, '20 min', 'Actif'],
        ['R002', 'Salade César', 'Entrées', 12.00, '', '', '', 1, '10 min', 'Actif'],
        ['R003', 'Saumon Grillé', 'Plats Principaux', 22.00, '', '', '', 1, '25 min', 'Actif'],
        ['R004', 'Tiramisu', 'Desserts', 8.00, '', '', '', 1, '30 min', 'Actif'],
        ['R005', 'Soupe du Jour', 'Entrées', 9.00, '', '', '', 1, '45 min', 'Actif']
    ]
    
    for i, data in enumerate(sample_data, 3):
        for j, val in enumerate(data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = FONTS['normal']
            cell.alignment = ALIGN_CENTER if j in [1, 4, 5, 6, 7, 8, 9] else ALIGN_LEFT
            cell.border = THIN_BORDER
            
            # Formules
            if j == 5:  # Coût matière
                cell.value = '=SUMIF(INGREDIENTS!A:A,A3,INGREDIENTS!F:F)'
            elif j == 6:  # Marge brute
                cell.value = '=D3-E3'
            elif j == 7:  # Taux marge
                cell.value = '=IFERROR(F3/D3*100,0)'
            
            if j in [4, 5, 6]:
                cell.number_format = '#,##0.00 €'
            elif j == 7:
                cell.number_format = '0.00%'
    
    # Ajuster largeurs
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 10
    
    create_table(ws, 'TableauRecettes', 2, 1, 100, 10, headers)
    
    return ws

def create_ingredients_sheet(wb):
    """Crée la feuille des ingrédients par recette"""
    ws = wb.create_sheet('INGREDIENTS')
    
    # Titre
    ws.merge_cells('A1:H1')
    title = ws.cell(row=1, column=1, value='🛒 DÉTAIL DES INGRÉDIENTS PAR RECETTE')
    title.font = Font(name='Calibri', size=16, bold=True, color=Color(rgb='FFFFFF'))
    title.fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    title.alignment = ALIGN_CENTER
    
    # En-têtes
    headers = ['ID Recette', 'Nom Plat', 'ID Ingrédient', 'Nom Ingrédient', 
               'Quantité', 'Unité', 'Coût Unitaire', 'Coût Total']
    create_header_row(ws, 2, headers, 1)
    
    # Données exemple
    sample_data = [
        ['R001', 'Burger Maison', 'P002', 'Boeuf Haché', 200, 'g', 0.012, ''],
        ['R001', 'Burger Maison', 'P010', 'Pain Burger', 1, 'unité', 0.50, ''],
        ['R001', 'Burger Maison', 'P011', 'Fromage', 50, 'g', 0.02, ''],
        ['R002', 'Salade César', 'P012', 'Laitue', 150, 'g', 0.008, ''],
        ['R002', 'Salade César', 'P013', 'Parmesan', 30, 'g', 0.03, ''],
        ['R003', 'Saumon Grillé', 'P005', 'Saumon', 250, 'g', 0.018, ''],
        ['R003', 'Saumon Grillé', 'P004', 'Huile Olive', 20, 'ml', 0.0085, '']
    ]
    
    for i, data in enumerate(sample_data, 3):
        for j, val in enumerate(data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = FONTS['normal']
            cell.alignment = ALIGN_CENTER if j in [1, 3, 5, 6, 7, 8] else ALIGN_LEFT
            cell.border = THIN_BORDER
            
            # Formule coût total
            if j == 8:
                cell.value = '=E3*G3'
                cell.number_format = '#,##0.00 €'
            
            if j == 7:
                cell.number_format = '#,##0.00 €'
    
    # Ajuster largeurs
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    
    create_table(ws, 'TableauIngredients', 2, 1, 200, 8, headers)
    
    return ws

def create_ventes_sheet(wb):
    """Crée la feuille de gestion des ventes"""
    ws = wb.create_sheet('VENTES')
    
    # Titre
    ws.merge_cells('A1:L1')
    title = ws.cell(row=1, column=1, value='💰 GESTION DES VENTES ET CHIFFRE D\'AFFAIRES')
    title.font = Font(name='Calibri', size=16, bold=True, color=Color(rgb='FFFFFF'))
    title.fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    title.alignment = ALIGN_CENTER
    
    # En-têtes
    headers = ['Date', 'ID Vente', 'Serveur', 'Table', 'ID Recette', 'Nom Plat', 
               'Prix Vente', 'Coût Matière', 'Marge', 'Quantité', 'Total HT', 'Statut']
    create_header_row(ws, 2, headers, 1)
    
    # Données exemple
    sample_data = [
        ['01/01/2024', 'V001', 'Marie', 'T05', 'R001', 'Burger Maison', 15.00, '', '', 2, '', 'Payé'],
        ['01/01/2024', 'V002', 'Pierre', 'T03', 'R002', 'Salade César', 12.00, '', '', 1, '', 'Payé'],
        ['01/01/2024', 'V003', 'Marie', 'T07', 'R003', 'Saumon Grillé', 22.00, '', '', 1, '', 'Payé'],
        ['02/01/2024', 'V004', 'Sophie', 'T02', 'R001', 'Burger Maison', 15.00, '', '', 3, '', 'Payé'],
        ['02/01/2024', 'V005', 'Pierre', 'T08', 'R004', 'Tiramisu', 8.00, '', '', 2, '', 'Payé']
    ]
    
    for i, data in enumerate(sample_data, 3):
        for j, val in enumerate(data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = FONTS['normal']
            cell.alignment = ALIGN_CENTER if j in [1, 2, 4, 5, 7, 8, 9, 10, 11, 12] else ALIGN_LEFT
            cell.border = THIN_BORDER
            
            # Formules
            if j == 8:  # Coût matière
                cell.value = '=LOOKUP(E3,RECETTES!A:A,RECETTES!E:E)'
            elif j == 9:  # Marge unitaire
                cell.value = '=G3-H3'
            elif j == 11:  # Total HT
                cell.value = '=G3*J3'
            
            if j in [7, 8, 9, 11]:
                cell.number_format = '#,##0.00 €'
    
    # Ajuster largeurs
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 10
    
    create_table(ws, 'TableauVentes', 2, 1, 500, 12, headers)
    
    return ws

def create_fournisseurs_sheet(wb):
    """Crée la feuille de gestion des fournisseurs"""
    ws = wb.create_sheet('FOURNISSEURS')
    
    # Titre
    ws.merge_cells('A1:K1')
    title = ws.cell(row=1, column=1, value='🚚 GESTION DES FOURNISSEURS ET COMMANDES')
    title.font = Font(name='Calibri', size=16, bold=True, color=Color(rgb='FFFFFF'))
    title.fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    title.alignment = ALIGN_CENTER
    
    # En-têtes
    headers = ['ID Fourn.', 'Nom Fournisseur', 'Contact', 'Téléphone', 'Email', 'Adresse', 
               'Statut Commande', 'Dernière Cmd', 'Montant Total', 'Note Qualité', 'Délai Livraison']
    create_header_row(ws, 2, headers, 1)
    
    # Données exemple
    sample_data = [
        ['F001', 'Fresh Market', 'Jean Dupont', '01 23 45 67 89', 'contact@freshmarket.fr', 'Paris', 'Actif', '01/01/2024', 1500.00, 4.5, '24h'],
        ['F002', 'Meat Pro', 'Marie Martin', '01 98 76 54 32', 'marie@meatpro.fr', 'Lyon', 'En attente', '28/12/2023', 2300.00, 4.8, '48h'],
        ['F003', 'Seafood Direct', 'Pierre Durand', '02 11 22 33 44', 'pierre@seafood.fr', 'Marseille', 'Actif', '30/12/2023', 1800.00, 4.2, '24h'],
        ['F004', 'Bio Légumes', 'Sophie Bernard', '03 55 66 77 88', 'sophie@biolegumes.fr', 'Bordeaux', 'Actif', '29/12/2023', 950.00, 4.9, '48h'],
        ['F005', 'Épices & Co', 'Luc Petit', '04 77 88 99 00', 'luc@epicesco.fr', 'Nice', 'Inactif', '15/12/2023', 450.00, 3.8, '72h']
    ]
    
    for i, data in enumerate(sample_data, 3):
        for j, val in enumerate(data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = FONTS['normal']
            cell.alignment = ALIGN_CENTER if j in [1, 4, 7, 8, 9, 10, 11] else ALIGN_LEFT
            cell.border = THIN_BORDER
            
            if j == 9:
                cell.number_format = '#,##0.00 €'
    
    # Ajuster largeurs
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 12
    
    create_table(ws, 'TableauFournisseurs', 2, 1, 100, 11, headers)
    
    return ws

def create_inventaires_sheet(wb):
    """Crée la feuille de gestion des inventaires"""
    ws = wb.create_sheet('INVENTAIRES')
    
    # Titre
    ws.merge_cells('A1:J1')
    title = ws.cell(row=1, column=1, value='📝 GESTION DES INVENTAIRES ET ÉCARTS')
    title.font = Font(name='Calibri', size=16, bold=True, color=Color(rgb='FFFFFF'))
    title.fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    title.alignment = ALIGN_CENTER
    
    # En-têtes
    headers = ['Date', 'ID Inventaire', 'ID Produit', 'Nom Produit', 'Stock Théorique', 
               'Stock Réel', 'Écart', 'Valeur Écart', 'Cause', 'Responsable']
    create_header_row(ws, 2, headers, 1)
    
    # Données exemple
    sample_data = [
        ['01/01/2024', 'INV001', 'P001', 'Tomates', 50, 48, '', '', 'Perte normale', 'Chef'],
        ['01/01/2024', 'INV001', 'P002', 'Boeuf Haché', 15, 14, '', '', 'Évaporation', 'Chef'],
        ['01/01/2024', 'INV001', 'P003', 'Riz Basmati', 100, 95, '', '', 'Erreur saisie', 'Magasinier'],
        ['15/01/2024', 'INV002', 'P004', 'Huile Olive', 20, 18, '', '', 'Casse', 'Chef'],
        ['15/01/2024', 'INV002', 'P005', 'Saumon', 8, 7, '', '', 'Perte qualité', 'Chef']
    ]
    
    for i, data in enumerate(sample_data, 3):
        for j, val in enumerate(data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = FONTS['normal']
            cell.alignment = ALIGN_CENTER if j in [1, 2, 3, 5, 6, 7, 8] else ALIGN_LEFT
            cell.border = THIN_BORDER
            
            # Formules
            if j == 7:  # Écart
                cell.value = '=F3-E3'
            elif j == 8:  # Valeur écart
                cell.value = '=G3*LOOKUP(C3,PRODUITS!A:A,PRODUITS!F:F)'
            
            if j == 8:
                cell.number_format = '#,##0.00 €'
    
    # Ajuster largeurs
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 15
    
    create_table(ws, 'TableauInventaires', 2, 1, 200, 10, headers)
    
    return ws

def create_tableaux_bord_sheet(wb):
    """Crée la feuille des tableaux de bord analytiques"""
    ws = wb.create_sheet('TABLEAUX_BORD')
    
    # Titre
    ws.merge_cells('A1:M1')
    title = ws.cell(row=1, column=1, value='📊 TABLEAUX DE BORD ANALYTIQUES')
    title.font = Font(name='Calibri', size=16, bold=True, color=Color(rgb='FFFFFF'))
    title.fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    title.alignment = ALIGN_CENTER
    
    # Section 1: Performance Ventes
    ws.merge_cells('A3:D3')
    section1 = ws.cell(row=3, column=1, value='📈 PERFORMANCE VENTES')
    section1.font = FONTS['subtitle']
    section1.alignment = ALIGN_CENTER
    
    kpi_ventes = [
        ('A4', 'CA Total', '=SUM(VENTES!K:K)'),
        ('C4', 'Nb Transactions', '=COUNTA(VENTES!A:A)-1'),
        ('A6', 'Panier Moyen', '=IFERROR(A4/C4,0)'),
        ('C6', 'Marge Totale', '=SUM(VENTES!I:I)')
    ]
    
    for cell_ref, label, formula in kpi_ventes:
        col = ord(cell_ref[0]) - 64
        row = int(cell_ref[1:])
        ws.cell(row=row, column=col, value=label).font = Font(name='Calibri', size=10, bold=True)
        val_cell = ws.cell(row=row, column=col+1, value=formula)
        val_cell.font = Font(name='Calibri', size=11, bold=True, color=Color(rgb=COLORS['success']))
        val_cell.number_format = '#,##0.00 €'
    
    # Section 2: Analyse Stocks
    ws.merge_cells('F3:I3')
    section2 = ws.cell(row=3, column=6, value='📦 ANALYSE STOCKS')
    section2.font = FONTS['subtitle']
    section2.alignment = ALIGN_CENTER
    
    kpi_stocks = [
        ('F4', 'Valeur Stock', '=SUMPRODUCT(PRODUITS!F:F,PRODUITS!G:G)'),
        ('H4', 'Produits Critiques', '=COUNTIF(PRODUITS!H:H,"⚠️ CRITIQUE")+COUNTIF(PRODUITS!H:H,"⚠️ RUPTURE")'),
        ('F6', 'Rotation Stock', '=SUM(VENTES!K:K)/F4'),
        ('H6', 'Couverture Jours', '=F4/(SUM(VENTES!K:K)/30)')
    ]
    
    for cell_ref, label, formula in kpi_stocks:
        col = ord(cell_ref[0]) - 64
        row = int(cell_ref[1:])
        ws.cell(row=row, column=col, value=label).font = Font(name='Calibri', size=10, bold=True)
        val_cell = ws.cell(row=row, column=col+1, value=formula)
        val_cell.font = Font(name='Calibri', size=11, bold=True, color=Color(rgb=COLORS['warning']))
        val_cell.number_format = '#,##0.00'
    
    # Section 3: Top Produits
    ws.merge_cells('A9:D9')
    section3 = ws.cell(row=9, column=1, value='🏆 TOP PRODUITS')
    section3.font = FONTS['subtitle']
    section3.alignment = ALIGN_CENTER
    
    headers_top = ['Rang', 'Produit', 'Ventes', 'Marge']
    create_header_row(ws, 10, headers_top, 1)
    
    # Section 4: Alertes
    ws.merge_cells('F9:I9')
    section4 = ws.cell(row=9, column=6, value='🔔 ALERTES EN COURS')
    section4.font = FONTS['subtitle']
    section4.alignment = ALIGN_CENTER
    
    alert_headers = ['Type', 'Nombre', 'Priorité']
    create_header_row(ws, 10, alert_headers, 6)
    
    alerts = [
        ('F11', 'Ruptures Stock', '=COUNTIF(PRODUITS!H:H,"⚠️ RUPTURE")', 'HAUTE'),
        ('F12', 'Stocks Critiques', '=COUNTIF(PRODUITS!H:H,"⚠️ CRITIQUE")', 'HAUTE'),
        ('F13', 'Cmds En Attente', '=COUNTIF(FOURNISSEURS!G:G,"En attente")', 'MOYENNE'),
        ('F14', 'Écarts Inventaire', '=COUNTIF(INVENTAIRES!G:G,"<>0")', 'BASSE')
    ]
    
    for cell_ref, label, formula, priority in alerts:
        row = int(cell_ref[1:])
        ws.cell(row=row, column=6, value=label).font = FONTS['normal']
        ws.cell(row=row, column=7, value=formula).font = Font(name='Calibri', size=10, bold=True)
        ws.cell(row=row, column=8, value=priority).font = FONTS['normal']
    
    # Instructions
    ws.merge_cells('A16:M16')
    instr = ws.cell(row=16, column=1, value='💡 Pour actualiser les données : Données → Actualiser tout | Pour créer des graphiques : Sélectionnez les données → Insertion → Graphique')
    instr.font = Font(name='Calibri', size=10, italic=True)
    instr.alignment = ALIGN_CENTER
    
    return ws

def create_parametres_sheet(wb):
    """Crée la feuille des paramètres"""
    ws = wb.create_sheet('PARAMETRES')
    
    # Titre
    ws.merge_cells('A1:E1')
    title = ws.cell(row=1, column=1, value='⚙️ PARAMÈTRES ET CONFIGURATION')
    title.font = Font(name='Calibri', size=16, bold=True, color=Color(rgb='FFFFFF'))
    title.fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    title.alignment = ALIGN_CENTER
    
    # Informations Restaurant (lignes 3-8)
    ws.merge_cells('A3:E3')
    section1 = ws.cell(row=3, column=1, value='Informations Restaurant')
    section1.font = FONTS['subtitle']
    section1.alignment = ALIGN_CENTER
    
    info_fields = ['Nom:', 'Adresse:', 'Téléphone:', 'Email:', 'SIRET:']
    info_values = ['Le Gourmet', '123 Rue de la Gastronomie', '01 23 45 67 89', 
                   'contact@restaurant.fr', '123456789']
    
    for i, field in enumerate(info_fields, 4):
        ws.cell(row=i, column=1, value=field).font = FONTS['normal']
        cell = ws.cell(row=i, column=2, value=info_values[i-4])
        cell.border = THIN_BORDER
    
    # Seuils d'Alerte (lignes 9-13)
    ws.merge_cells('A9:E9')
    section2 = ws.cell(row=9, column=1, value='Seuils d\'Alerte')
    section2.font = FONTS['subtitle']
    section2.alignment = ALIGN_CENTER
    
    seuil_fields = ['Stock Minimum Par Défaut:', 'Stock Sécurité Par Défaut:', 
                    'Jours Avant Inventaire:', 'Seuil Marge Minimum (%):']
    seuil_values = ['20', '10', '7', '30']
    
    for i, field in enumerate(seuil_fields, 10):
        ws.cell(row=i, column=1, value=field).font = FONTS['normal']
        cell = ws.cell(row=i, column=2, value=seuil_values[i-10])
        cell.border = THIN_BORDER
    
    # Options Système (lignes 14-18)
    ws.merge_cells('A14:E14')
    section3 = ws.cell(row=14, column=1, value='Options Système')
    section3.font = FONTS['subtitle']
    section3.alignment = ALIGN_CENTER
    
    option_fields = ['Devise:', 'Format Date:', 'Fuseau Horaire:', 'Langue:']
    option_values = ['€', 'JJ/MM/AAAA', 'Europe/Paris', 'Français']
    
    for i, field in enumerate(option_fields, 15):
        ws.cell(row=i, column=1, value=field).font = FONTS['normal']
        cell = ws.cell(row=i, column=2, value=option_values[i-15])
        cell.border = THIN_BORDER
    
    # Ajuster largeurs
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    return ws

def create_aide_sheet(wb):
    """Crée la feuille d'aide"""
    ws = wb.create_sheet('AIDE')
    
    # Titre
    ws.merge_cells('A1:F1')
    title = ws.cell(row=1, column=1, value='❓ GUIDE D\'UTILISATION ET AIDE')
    title.font = Font(name='Calibri', size=16, bold=True, color=Color(rgb='FFFFFF'))
    title.fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    title.alignment = ALIGN_CENTER
    
    # Contenu
    content = [
        ('A3', '🚀 PRISE EN MAIN RAPIDE', '''1. Commencez par remplir la feuille CATÉGORIES
2. Ajoutez vos PRODUITS avec leurs informations
3. Configurez vos FOURNISSEURS
4. Créez vos RECETTES et ajoutez les INGRÉDIENTS
5. Enregistrez les MOUVEMENTS DE STOCK
6. Saisissez les VENTES quotidiennes
7. Consultez le DASHBOARD pour les indicateurs'''),
        
        ('A8', '📊 FORMULES CLÉS', '''• Statut Stock: =IF(G2=0,"⚠️ RUPTURE",IF(G2<D2,"⚠️ CRITIQUE",IF(G2<E2,"⚡ ATTENTION","✓ OK")))
• Coût Matière Recette: =SUMIF(INGREDIENTS!A:A,A3,INGREDIENTS!F:F)
• Marge Brute: =Prix_Vente-Coût_Matière
• Taux Marge: =Marge/Prix_Vente*100'''),
        
        ('A13', '🔧 MAINTENANCE', '''• Actualisez les données: Données → Actualiser tout
• Sauvegardez régulièrement: Fichier → Enregistrer sous
• Protégez les feuilles: Révision → Protéger la feuille
• Vérifiez les formules après modifications'''),
        
        ('A18', '⚠️ BONNES PRATIQUES', '''• Faites des inventaires réguliers (hebdomadaires)
• Vérifiez les alertes de stock quotidiennement
• Analysez les marges par plat mensuellement
• Sauvegardez avant toute modification structurelle''')
    ]
    
    for start_cell, section_title, text in content:
        row = int(start_cell[1:])
        ws.merge_cells(f'A{row}:F{row}')
        section = ws.cell(row=row, column=1, value=section_title)
        section.font = FONTS['subtitle']
        section.alignment = ALIGN_LEFT
        
        text_cell = ws.cell(row=row+1, column=1, value=text)
        text_cell.font = Font(name='Calibri', size=10)
        text_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[row+1].height = 100
    
    return ws

def main():
    """Fonction principale de création du fichier Excel"""
    print("🍽️ Création de l'application Excel de gestion de restaurant...")
    
    # Créer le classeur
    wb = Workbook()
    
    # Supprimer la feuille par défaut
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Créer toutes les feuilles
    sheets = [
        ('ACCUEIL', create_home_sheet),
        ('DASHBOARD', create_dashboard_sheet),
        ('PRODUITS', create_products_sheet),
        ('CATEGORIES', create_categories_sheet),
        ('STOCKS', create_stocks_sheet),
        ('RECETTES', create_recettes_sheet),
        ('INGREDIENTS', create_ingredients_sheet),
        ('VENTES', create_ventes_sheet),
        ('FOURNISSEURS', create_fournisseurs_sheet),
        ('INVENTAIRES', create_inventaires_sheet),
        ('TABLEAUX_BORD', create_tableaux_bord_sheet),
        ('PARAMETRES', create_parametres_sheet),
        ('AIDE', create_aide_sheet)
    ]
    
    for sheet_name, create_func in sheets:
        print(f"  📄 Création de la feuille: {sheet_name}")
        create_func(wb)
    
    # Enregistrer le fichier
    output_file = '/workspace/GestionRestaurant.xlsx'
    wb.save(output_file)
    print(f"\n✅ Fichier Excel créé avec succès: {output_file}")
    print(f"📊 Nombre de feuilles: {len(wb.sheetnames)}")
    print(f"📋 Feuilles créées: {', '.join(wb.sheetnames)}")
    
    return output_file

if __name__ == '__main__':
    main()
